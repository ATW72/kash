import os
import sys
import sqlite3
import datetime as dt
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
from datetime import datetime, timedelta
from flask import Flask, jsonify, request, session, redirect, url_for, render_template, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
import io
import csv
import json

sys.path.insert(0, os.path.dirname(__file__))
from config.settings import Config
from utils.db import get_db_connection
from utils.auth import login_required, admin_required
from utils.notifications import init_mail, run_daily_notifications, send_email, build_bill_alert_email, build_budget_alert_email, build_welcome_email
import secrets
import string
import math
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
from authlib.integrations.flask_client import OAuth
from werkzeug.middleware.proxy_fix import ProxyFix

# ── Google OAuth Config ───────────────────────────────────────────────────────


app = Flask(__name__, static_folder='static', static_url_path='/static')
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
import pathlib
UPLOAD_FOLDER = pathlib.Path(os.environ.get('APP_UPLOAD_FOLDER', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'receipts')))
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)
ALLOWED_EXTENSIONS = {'png','jpg','jpeg','gif','webp','pdf'}
app.secret_key = Config.SECRET_KEY
app.config['SESSION_COOKIE_SECURE'] = Config.SESSION_COOKIE_SECURE
app.config['SESSION_COOKIE_HTTPONLY'] = Config.SESSION_COOKIE_HTTPONLY
app.config['SESSION_COOKIE_SAMESITE'] = Config.SESSION_COOKIE_SAMESITE
app.config['PERMANENT_SESSION_LIFETIME'] = Config.PERMANENT_SESSION_LIFETIME
CORS(app)

oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=Config.GOOGLE_CLIENT_ID,
    client_secret=Config.GOOGLE_CLIENT_SECRET,
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={
        'scope': 'openid email profile',
    }
)


# ── Database ──────────────────────────────────────────────────────────────────

def init_db():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("PRAGMA foreign_keys = ON")

    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        is_admin INTEGER NOT NULL DEFAULT 0,
        display_name TEXT NOT NULL DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    # Migration: add display_name to existing databases
    try:
        c.execute("ALTER TABLE users ADD COLUMN display_name TEXT NOT NULL DEFAULT ''")
        conn.commit()
    except Exception:
        pass  # Column already exists, safe to ignore

    # Migration: must_change_password flag
    try:
        c.execute("ALTER TABLE users ADD COLUMN must_change_password INTEGER DEFAULT 0")
        conn.commit()
    except Exception:
        pass

    # Migrations: email/notification columns on users
    for col, defn in [
        ("email",           "TEXT DEFAULT ''"),
        ("notify_bills",    "INTEGER DEFAULT 1"),
        ("notify_budgets",  "INTEGER DEFAULT 1"),
        ("notify_monthly",  "INTEGER DEFAULT 1"),
        ("two_factor_method", "TEXT DEFAULT 'none'"),
        ("totp_secret",       "TEXT DEFAULT ''"),
        ("google_id",       "TEXT DEFAULT ''"),
        ("avatar_url",      "TEXT DEFAULT ''"),
    ]:
        try:
            c.execute(f"ALTER TABLE users ADD COLUMN {col} {defn}")
            conn.commit()
        except Exception:
            pass

    # Migrations: receipt_filename on expenses

    # Migrations: currency columns on expenses and income
    for tbl in ['expenses', 'income']:
        for col, defn in [("currency", "TEXT DEFAULT 'USD'"), ("original_amount", "REAL")]:
            try:
                c.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {defn}")
                conn.commit()
            except Exception:
                pass

    # Migrations: is_recurring and frequency on bills
    for col, defn in [("is_recurring", "INTEGER DEFAULT 0"), ("frequency", "TEXT DEFAULT 'monthly'")]:
        try:
            c.execute(f"ALTER TABLE bills ADD COLUMN {col} {defn}")
            conn.commit()
        except Exception:
            pass

    # Migration: is_business column for data isolation
    for tbl in ['expenses', 'income', 'budgets', 'bills', 'recurring_transactions']:
        try:
            c.execute(f"ALTER TABLE {tbl} ADD COLUMN is_business INTEGER DEFAULT 0")
            conn.commit()
        except Exception:
            pass

    # Migration: category type (personal vs business)
    try:
        c.execute("ALTER TABLE categories ADD COLUMN type TEXT DEFAULT 'personal'")
        conn.commit()
    except Exception:
        pass

    # Migrations: rollover on budgets



    c.execute("""CREATE TABLE IF NOT EXISTS categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        type TEXT DEFAULT 'personal')""")

    c.execute("""CREATE TABLE IF NOT EXISTS credit_cards (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner TEXT NOT NULL,
        card_name TEXT NOT NULL,
        credit_limit REAL,
        manual_balance REAL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    c.execute("""CREATE TABLE IF NOT EXISTS expenses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        category TEXT NOT NULL,
        description TEXT NOT NULL,
        amount REAL NOT NULL,
        paid_by TEXT NOT NULL,
        payment_method TEXT,
        credit_card_id INTEGER,
        notes TEXT,
        receipt_filename TEXT,
        is_business INTEGER DEFAULT 0,
        owner TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (credit_card_id) REFERENCES credit_cards(id) ON DELETE SET NULL)""")

    c.execute("""CREATE TABLE IF NOT EXISTS income (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        source TEXT NOT NULL,
        description TEXT NOT NULL,
        amount REAL NOT NULL,
        received_by TEXT NOT NULL,
        notes TEXT,
        is_business INTEGER DEFAULT 0,
        owner TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    c.execute("""CREATE TABLE IF NOT EXISTS budgets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category TEXT NOT NULL,
        amount REAL NOT NULL,
        month INTEGER NOT NULL,
        year INTEGER NOT NULL,
        rollover REAL NOT NULL DEFAULT 0,
        is_business INTEGER DEFAULT 0,
        owner TEXT,
        UNIQUE(category, month, year))""")

    c.execute("""CREATE TABLE IF NOT EXISTS savings_goals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        target_amount REAL NOT NULL,
        target_date TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    # Migration: create savings_goals if not exists (handled by CREATE IF NOT EXISTS above)

    c.execute("""CREATE TABLE IF NOT EXISTS bills (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        amount REAL NOT NULL,
        due_date TEXT NOT NULL,
        category TEXT,
        is_paid INTEGER DEFAULT 0,
        is_recurring INTEGER DEFAULT 0,
        frequency TEXT DEFAULT 'monthly',
        notes TEXT,
        is_business INTEGER DEFAULT 0,
        owner TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    c.execute("""CREATE TABLE IF NOT EXISTS recurring_transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT NOT NULL CHECK(type IN ('expense','income')),
        frequency TEXT NOT NULL,
        next_date TEXT NOT NULL,
        description TEXT NOT NULL,
        amount REAL NOT NULL,
        category TEXT,
        source TEXT,
        person TEXT NOT NULL,
        payment_method TEXT,
        credit_card_id INTEGER,
        is_active INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (credit_card_id) REFERENCES credit_cards(id) ON DELETE SET NULL)""")

    # (expense_splits, widget_prefs, onboarding defined below with other new tables)
    c.execute("""CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action TEXT NOT NULL,
        table_name TEXT NOT NULL,
        record_id INTEGER,
        username TEXT NOT NULL,
        changes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    c.execute("""CREATE TABLE IF NOT EXISTS currencies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT NOT NULL UNIQUE,
        name TEXT NOT NULL,
        symbol TEXT NOT NULL,
        rate_to_home REAL NOT NULL DEFAULT 1.0,
        is_home INTEGER NOT NULL DEFAULT 0,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    # Default home currency
    existing_cur = c.execute("SELECT COUNT(*) FROM currencies").fetchone()[0]
    if existing_cur == 0:
        c.execute("INSERT INTO currencies (code,name,symbol,rate_to_home,is_home) VALUES ('USD','US Dollar','$',1.0,1)")

    # Add currency columns to expenses and income if missing
    for tbl in ['expenses','income']:
        for col, defn in [('currency',"TEXT DEFAULT 'USD'"), ('original_amount','REAL')]:
            try:
                c.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {defn}")
                conn.commit()
            except Exception:
                pass

    c.execute("""CREATE TABLE IF NOT EXISTS expense_splits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        expense_id INTEGER NOT NULL,
        username TEXT NOT NULL,
        split_amount REAL NOT NULL,
        split_pct REAL,
        is_settled INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (expense_id) REFERENCES expenses(id) ON DELETE CASCADE)""")

    c.execute("""CREATE TABLE IF NOT EXISTS widget_prefs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL,
        widget_id TEXT NOT NULL,
        is_visible INTEGER NOT NULL DEFAULT 1,
        sort_order INTEGER NOT NULL DEFAULT 0,
        UNIQUE(username, widget_id))""")

    c.execute("""CREATE TABLE IF NOT EXISTS onboarding (
        username TEXT PRIMARY KEY,
        completed INTEGER DEFAULT 0,
        step INTEGER DEFAULT 0,
        completed_at TIMESTAMP)""")


    c.execute("""CREATE TABLE IF NOT EXISTS sharing (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        table_name TEXT NOT NULL,
        record_id INTEGER NOT NULL,
        owner TEXT NOT NULL,
        shared_with TEXT NOT NULL,
        can_edit INTEGER NOT NULL DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(table_name, record_id, shared_with))""")

    # Add owner column to all shareable tables
    for tbl, default_owner in [
        ('expenses',      ''),
        ('income',        ''),
        ('credit_cards',  ''),
        ('bills',         ''),
        ('budgets',       ''),
        ('savings_goals', ''),
        ('recurring_transactions', ''),
    ]:
        try:
            c.execute(f"ALTER TABLE {tbl} ADD COLUMN owner TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass

    # Create indexes for sharing lookups
    for sql in [
        "CREATE INDEX IF NOT EXISTS idx_sharing_table_record ON sharing(table_name, record_id)",
        "CREATE INDEX IF NOT EXISTS idx_sharing_shared_with ON sharing(shared_with)",
        "CREATE INDEX IF NOT EXISTS idx_expenses_owner ON expenses(owner)",
        "CREATE INDEX IF NOT EXISTS idx_income_owner ON income(owner)",
        "CREATE INDEX IF NOT EXISTS idx_bills_owner ON bills(owner)",
        "CREATE INDEX IF NOT EXISTS idx_budgets_owner ON budgets(owner)",
    ]:
        try:
            c.execute(sql)
        except Exception:
            pass

    for sql in [
        "CREATE INDEX IF NOT EXISTS idx_expenses_date ON expenses(date)",
        "CREATE INDEX IF NOT EXISTS idx_expenses_category ON expenses(category)",
        "CREATE INDEX IF NOT EXISTS idx_expenses_paid_by ON expenses(paid_by)",
        "CREATE INDEX IF NOT EXISTS idx_income_date ON income(date)",
        "CREATE INDEX IF NOT EXISTS idx_income_source ON income(source)",
    ]:
        c.execute(sql)

    # New Tables for Loans and CC Payments
    c.execute("""CREATE TABLE IF NOT EXISTS loans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner TEXT NOT NULL,
        loan_name TEXT NOT NULL,
        total_amount REAL NOT NULL,
        monthly_payment REAL,
        next_due_date TEXT,
        category TEXT,
        is_active INTEGER DEFAULT 1,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    # Migrations for loans
    for col, defn in [("interest_rate", "REAL DEFAULT 0")]:
        try:
            c.execute(f"ALTER TABLE loans ADD COLUMN {col} {defn}")
            conn.commit()
        except Exception:
            pass

    c.execute("""CREATE TABLE IF NOT EXISTS loan_payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        loan_id INTEGER NOT NULL,
        amount REAL NOT NULL,
        date TEXT NOT NULL,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (loan_id) REFERENCES loans(id) ON DELETE CASCADE)""")

    c.execute("""CREATE TABLE IF NOT EXISTS cc_payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        credit_card_id INTEGER NOT NULL,
        amount REAL NOT NULL,
        date TEXT NOT NULL,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (credit_card_id) REFERENCES credit_cards(id) ON DELETE CASCADE)""")

    # Migrations for credit_cards
    for col, defn in [("starting_balance", "REAL DEFAULT 0"), ("interest_rate", "REAL DEFAULT 0")]:
        try:
            c.execute(f"ALTER TABLE credit_cards ADD COLUMN {col} {defn}")
            conn.commit()
        except Exception:
            pass



    # Default categories
    for cat in ['Groceries','Dining Out','Transportation','Utilities','Entertainment',
                'Healthcare','Shopping','Mortgage','Child Support','Education','Other']:
        c.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (cat,))

    # Business categories
    for cat in ['Business Taxes', 'Payroll', 'Software Subscriptions', 'Office Supplies', 
                'Marketing', 'Inventory', 'Business Travel', 'Legal/Professional']:
        c.execute("INSERT OR IGNORE INTO categories (name, type) VALUES (?, 'business')", (cat,))

    # Bootstrap admin
    admin_user = os.environ.get('APP_LOGIN_USERNAME')
    admin_pass = os.environ.get('APP_LOGIN_PASSWORD')
    if admin_user and admin_pass:
        phash = generate_password_hash(admin_pass, method='pbkdf2:sha256')
        c.execute("INSERT OR IGNORE INTO users (username, password_hash, is_admin) VALUES (?,?,1)",
                  (admin_user, phash))
    else:
        # Fallback to generate forced admin replacement
        phash = generate_password_hash(secrets.token_urlsafe(16), method='pbkdf2:sha256')
        c.execute("INSERT OR IGNORE INTO users (username, password_hash, is_admin, must_change_password) VALUES (?,?,1,1)",
                  ('admin', phash))

    # Invoices table for business invoice generation
    c.execute("""CREATE TABLE IF NOT EXISTS invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_number TEXT NOT NULL,
        client_name TEXT NOT NULL,
        client_email TEXT DEFAULT '',
        issue_date TEXT NOT NULL,
        due_date TEXT NOT NULL,
        items TEXT NOT NULL,
        notes TEXT DEFAULT '',
        tax_rate REAL DEFAULT 0,
        status TEXT DEFAULT 'draft',
        owner TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    conn.commit()
    conn.close()

def get_credit_card_balance(card_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT manual_balance, starting_balance FROM credit_cards WHERE id = ?", (card_id,))
    row = c.fetchone()
    if row and row[0] is not None:
        conn.close()
        return float(row[0])
    
    starting_bal = float(row[1]) if row and row[1] else 0.0
    
    # Sum expenses
    c.execute("SELECT COALESCE(SUM(amount),0.0) FROM expenses WHERE credit_card_id = ?", (card_id,))
    expenses_sum = float(c.fetchone()[0])
    
    # Sum payments
    c.execute("SELECT COALESCE(SUM(amount),0.0) FROM cc_payments WHERE credit_card_id = ?", (card_id,))
    payments_sum = float(c.fetchone()[0])
    
    conn.close()
    return starting_bal + expenses_sum - payments_sum


def log_audit(conn, action, table_name, record_id, changes=None):
    """Log an audit trail entry. action = CREATE | UPDATE | DELETE"""
    try:
        username = 'system'
        try:
            username = session.get('username', 'system')
        except Exception:
            pass
        conn.execute(
            "INSERT INTO audit_log (action,table_name,record_id,username,changes) VALUES (?,?,?,?,?)",
            (action, table_name, record_id, username,
             json.dumps(changes, default=str) if changes else None))
    except Exception:
        pass  # Never let audit logging break the main operation


def get_visible_clause(table_name, username, alias=''):
    """Returns a WHERE clause fragment and params that filters to records
    owned by username OR shared with username. Admins see all."""
    tbl = f"{alias}." if alias else ""
    if session.get('is_admin'):
        return "1=1", []
    # Only records owned by the user or explicitly shared are visible.
    clause = f"({tbl}owner = ? OR id IN (SELECT record_id FROM sharing WHERE table_name=? AND shared_with=?))"
    params = [username, table_name, username]
    return clause, params

def is_owner_or_shared(conn, table_name, record_id, username):
    """Check if user owns a record or has it shared with them."""
    row = conn.execute(f"SELECT owner FROM {table_name} WHERE id=?", (record_id,)).fetchone()
    if not row:
        return False, False  # not found
    owner = row['owner'] or ''
    if owner == '' or owner == username:
        return True, True  # owner
    shared = conn.execute(
        "SELECT can_edit FROM sharing WHERE table_name=? AND record_id=? AND shared_with=?",
        (table_name, record_id, username)).fetchone()
    if shared:
        return True, bool(shared['can_edit'])  # shared: (can_see, can_edit)
    return False, False


def validate_expense(data):
    for f in ['date','category','description','amount','paid_by']:
        if f not in data or not str(data[f]).strip():
            return False, f"Missing required field: {f}"
    try:
        amt = float(data['amount'])
        if amt <= 0:
            return False, "Amount must be greater than zero"
    except (ValueError, TypeError):
        return False, "Invalid amount value"
    try:
        datetime.strptime(data['date'], '%Y-%m-%d')
    except ValueError:
        return False, "Invalid date format. Use YYYY-MM-DD"
    if not str(data['paid_by']).strip():
        return False, "paid_by is required"
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM categories WHERE name = ?", (data['category'],))
    if c.fetchone()[0] == 0:
        conn.close()
        return False, "Invalid category"
    conn.close()
    return True, None

def validate_income(data):
    for f in ['date','source','description','amount','received_by']:
        if f not in data or not str(data[f]).strip():
            return False, f"Missing required field: {f}"
    try:
        amt = float(data['amount'])
        if amt <= 0:
            return False, "Amount must be greater than zero"
    except (ValueError, TypeError):
        return False, "Invalid amount value"
    try:
        datetime.strptime(data['date'], '%Y-%m-%d')
    except ValueError:
        return False, "Invalid date format"
    return True, None

# ── Page Routes ───────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    from flask import make_response
    resp = make_response(render_template('index.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

# ── Auth API ──────────────────────────────────────────────────────────────────

@app.route('/login', methods=['POST'])
def login():
    data = request.json or {}
    username = data.get('username','').strip()
    password = data.get('password','')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT id,username,password_hash,is_admin,display_name,must_change_password,two_factor_method,totp_secret,email FROM users WHERE username=?", (username,))
    user = c.fetchone()
    conn.close()
    if not user or not check_password_hash(user[2], password):
        return jsonify({'error': 'Invalid username or password'}), 401
        
    uid, uname, phash, is_admin, display_name, must_change_pw, two_factor_method, totp_secret, email = user
    two_factor_method = two_factor_method or 'none'
    code = data.get('code', '').strip()

    if two_factor_method != 'none':
        if not code:
            # First pass: valid credentials, but 2FA is required.
            if two_factor_method == 'email':
                import random
                otp = f"{random.randint(100000, 999999)}"
                session['email_otp'] = otp
                session['email_otp_username'] = uname
                if email:
                    from utils.notifications import send_email
                    html = f"<h2>Login Verification Code</h2><p>Your correct login code is: <strong>{otp}</strong></p><p>This code expires shortly.</p>"
                    send_email(email, '✅ Kash: Login Verification Code', html)
            return jsonify({'success': False, 'status': '2fa_required', 'method': two_factor_method, 'username': uname}), 200
        else:
            # Second pass: verify the code
            if two_factor_method == 'app':
                import pyotp
                if not totp_secret:
                    return jsonify({'error': '2FA secret not found. Check server DB.'}), 500
                totp = pyotp.TOTP(totp_secret)
                if not totp.verify(code):
                    return jsonify({'error': 'Invalid Authenticator code.'}), 401
            elif two_factor_method == 'email':
                if code != str(session.get('email_otp')) or uname != session.get('email_otp_username'):
                    return jsonify({'error': 'Invalid or expired email code.'}), 401
                session.pop('email_otp', None)
                session.pop('email_otp_username', None)

    session.permanent = True
    session['user_id'] = uid
    session['username'] = uname
    session['is_admin'] = bool(is_admin)
    session['display_name'] = display_name or ''
    session['must_change_password'] = bool(must_change_pw)
    return jsonify({'success': True, 'user': {'id': uid, 'username': uname, 'is_admin': bool(is_admin), 'display_name': display_name or '', 'must_change_password': bool(must_change_pw)}}), 200

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out successfully'}), 200

@app.route('/login/google')
def login_google():
    if 'user_id' in session:
        return redirect(url_for('index'))
    redirect_uri = url_for('auth_google', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route('/auth/google')
def auth_google():
    token = google.authorize_access_token()
    user_info = token.get('userinfo')
    
    email = user_info.get('email')
    google_id = user_info.get('id')
    display_name = user_info.get('name', '')
    avatar_url = user_info.get('picture', '')

    conn = get_db_connection()
    c = conn.cursor()
    
    c.execute("SELECT id, username, is_admin, must_change_password FROM users WHERE google_id=?", (google_id,))
    user = c.fetchone()
    
    if not user and email:
        c.execute("SELECT id, username, is_admin, must_change_password FROM users WHERE email=?", (email,))
        user = c.fetchone()
        if user:
            c.execute("UPDATE users SET google_id=?, avatar_url=? WHERE id=?", (google_id, avatar_url, user['id']))
            conn.commit()
            
    if not user:
        conn.close()
        return redirect(url_for('login', error='not_found'))

    conn.close()
    
    session.permanent = True
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['is_admin'] = bool(user['is_admin'])
    session['display_name'] = display_name
    session['must_change_password'] = False
    
    return redirect(url_for('index'))

@app.route('/api/auth/me')
@login_required
def get_current_user():
    return jsonify({'user': {'id': session.get('user_id'), 'username': session.get('username'), 'is_admin': session.get('is_admin', False), 'display_name': session.get('display_name', ''), 'must_change_password': session.get('must_change_password', False)}}), 200

# ── Categories ────────────────────────────────────────────────────────────────

@app.route('/api/categories')
@login_required
def get_categories():
    ctype = request.args.get('type', 'personal')
    conn = get_db_connection()
    q = "SELECT name FROM categories WHERE type = ? ORDER BY name"
    cats = [row['name'] for row in conn.execute(q, (ctype,)).fetchall()]
    conn.close()
    return jsonify({'data': cats}), 200

# ── Expenses ──────────────────────────────────────────────────────────────────

@app.route('/api/expenses', methods=['GET'])
@login_required
def get_expenses():
    start     = request.args.get('start_date')
    end       = request.args.get('end_date')
    category  = request.args.get('category')
    mine_only = request.args.get('mine_only') == '1'
    is_business = request.args.get('is_business', '0')
    username = session['username']
    conn = get_db_connection()
    if mine_only:
        q = "SELECT * FROM expenses WHERE (owner=? OR owner='')"
        params = [username]
    else:
        vis_clause, vis_params = get_visible_clause('expenses', username)
        q = f"SELECT * FROM expenses WHERE {vis_clause}"
        params = vis_params
    
    q += " AND is_business = ?"; params.append(int(is_business))
    if start:    q += " AND date >= ?";     params.append(start)
    if end:      q += " AND date <= ?";     params.append(end)
    if category: q += " AND category = ?"; params.append(category)
    q += " ORDER BY date DESC, created_at DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify({'data': [dict(r) for r in rows]}), 200

@app.route('/api/expenses', methods=['POST'])
@login_required
def create_expense():
    data = request.json or {}
    ok, err = validate_expense(data)
    if not ok:
        return jsonify({'error': err}), 400
    try:
        conn = get_db_connection()

        # Duplicate detection — warn if same description+amount within 3 days
        if not data.get('force'):
            from_date = datetime.strptime(data['date'], '%Y-%m-%d')
            date_range = [(from_date - __import__('datetime').timedelta(days=3)).strftime('%Y-%m-%d'),
                          (from_date + __import__('datetime').timedelta(days=3)).strftime('%Y-%m-%d')]
            dup = conn.execute(
                """SELECT id, date, description FROM expenses
                   WHERE description=? AND amount=? AND date BETWEEN ? AND ?""",
                (data['description'], float(data['amount']), date_range[0], date_range[1])
            ).fetchone()
            if dup:
                conn.close()
                return jsonify({'duplicate': True, 'existing': dict(dup),
                                'message': f'Similar expense exists on {dup["date"]}. Add anyway?'}), 409

        # Convert currency if not home currency
        amount_home = float(data['amount'])
        currency = data.get('currency', 'USD')
        original_amount = float(data['amount'])
        if currency:
            rate_row = conn.execute("SELECT rate_to_home FROM currencies WHERE code=?", (currency,)).fetchone()
            if rate_row and rate_row['rate_to_home']:
                amount_home = round(original_amount / rate_row['rate_to_home'], 2)

        c = conn.cursor()
        c.execute("""INSERT INTO expenses (date,category,description,amount,paid_by,payment_method,
                     credit_card_id,notes,currency,original_amount,owner,is_business)
                     VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                  (data['date'], data['category'], data['description'], amount_home,
                   data['paid_by'], data.get('payment_method'), data.get('credit_card_id'),
                   data.get('notes',''), currency, original_amount, session['username'],
                   int(data.get('is_business', 0))))
        eid = c.lastrowid
        log_audit(conn, 'CREATE', 'expenses', eid, {
            'description': data['description'], 'amount': amount_home,
            'date': data['date'], 'category': data['category']})
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': eid, 'message': 'Expense created successfully'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/expenses/<int:eid>', methods=['PUT'])
@login_required
def update_expense(eid):
    data = request.json or {}
    ok, err = validate_expense(data)
    if not ok:
        return jsonify({'error': err}), 400
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'expenses', eid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Expense not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        c = conn.cursor()
        old = conn.execute("SELECT * FROM expenses WHERE id=?", (eid,)).fetchone()
        c.execute("""UPDATE expenses SET date=?,category=?,description=?,amount=?,
                     paid_by=?,payment_method=?,credit_card_id=?,notes=?,is_business=? WHERE id=?""",
                  (data['date'], data['category'], data['description'], float(data['amount']),
                   data['paid_by'], data.get('payment_method'), data.get('credit_card_id'),
                   data.get('notes',''), int(data.get('is_business', 0)), eid))
        log_audit(conn, 'UPDATE', 'expenses', eid, {
            'before': {'description': old['description'], 'amount': old['amount'], 'date': old['date'], 'is_business': old['is_business']} if old else {},
            'after':  {'description': data['description'], 'amount': float(data['amount']), 'date': data['date'], 'is_business': int(data.get('is_business', 0))}})
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Expense updated'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/expenses/<int:eid>', methods=['DELETE'])
@login_required
def delete_expense(eid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'expenses', eid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Expense not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        c = conn.cursor()
        row = conn.execute("SELECT * FROM expenses WHERE id=?", (eid,)).fetchone()
        c.execute("DELETE FROM expenses WHERE id=?", (eid,))
        log_audit(conn, 'DELETE', 'expenses', eid, {
            'description': row['description'] if row else '', 'amount': row['amount'] if row else 0,
            'date': row['date'] if row else ''})
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Expense deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Income ────────────────────────────────────────────────────────────────────

@app.route('/api/income', methods=['GET'])
@login_required
def get_income():
    start     = request.args.get('start_date')
    end       = request.args.get('end_date')
    mine_only = request.args.get('mine_only') == '1'
    is_business = request.args.get('is_business', '0')
    username  = session['username']
    conn = get_db_connection()
    if mine_only:
        q = "SELECT * FROM income WHERE (owner=? OR owner='')"; params = [username]
    else:
        vis_clause, params = get_visible_clause('income', username)
        q = f"SELECT * FROM income WHERE {vis_clause}"
    
    q += " AND is_business = ?"; params.append(int(is_business))
    if start: q += " AND date >= ?"; params.append(start)
    if end:   q += " AND date <= ?"; params.append(end)
    q += " ORDER BY date DESC, created_at DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify({'data': [dict(r) for r in rows]}), 200

@app.route('/api/income', methods=['POST'])
@login_required
def create_income():
    data = request.json or {}
    ok, err = validate_income(data)
    if not ok:
        return jsonify({'error': err}), 400
    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("INSERT INTO income (date,source,description,amount,received_by,notes,owner,is_business) VALUES (?,?,?,?,?,?,?,?)",
                  (data['date'], data['source'], data['description'], float(data['amount']),
                   data['received_by'], data.get('notes',''), session['username'], int(data.get('is_business', 0))))
        iid = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': iid, 'message': 'Income created successfully'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/income/<int:iid>', methods=['PUT'])
@login_required
def update_income(iid):
    data = request.json or {}
    ok, err = validate_income(data)
    if not ok:
        return jsonify({'error': err}), 400
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'income', iid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Income not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        c = conn.cursor()
        c.execute("UPDATE income SET date=?,source=?,description=?,amount=?,received_by=?,notes=?,is_business=? WHERE id=?",
                  (data['date'], data['source'], data['description'], float(data['amount']),
                   data['received_by'], data.get('notes',''), int(data.get('is_business', 0)), iid))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/income/<int:iid>', methods=['DELETE'])
@login_required
def delete_income(iid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'income', iid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Income not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403
            
        c = conn.cursor()
        c.execute("DELETE FROM income WHERE id=?", (iid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Income deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    if c.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Income not found'}), 404
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

# ── Credit Cards ──────────────────────────────────────────────────────────────

@app.route('/api/credit-cards', methods=['GET'])
@login_required
def get_credit_cards():
    conn = get_db_connection()
    vis_clause, vis_params = get_visible_clause('credit_cards', session['username'])
    cards = [dict(r) for r in conn.execute(f"SELECT * FROM credit_cards WHERE {vis_clause} ORDER BY owner, card_name", vis_params).fetchall()]
    conn.close()
    for card in cards:
        card['balance'] = round(get_credit_card_balance(card['id']), 2)
        if card['credit_limit']:
            card['utilization'] = round(card['balance'] / card['credit_limit'] * 100, 1)
        else:
            card['utilization'] = None
    return jsonify({'data': cards}), 200

@app.route('/api/credit-cards', methods=['POST'])
@login_required
def create_credit_card():
    data = request.json or {}
    if not data.get('card_name'):
        return jsonify({'error': 'Card name required'}), 400
    conn = get_db_connection()
    c = conn.cursor()
    # Always set owner to current user
    c.execute("INSERT INTO credit_cards (owner,card_name,credit_limit,starting_balance,interest_rate) VALUES (?,?,?,?,?)",
              (session['username'], data['card_name'], data.get('credit_limit'), data.get('starting_balance', 0), data.get('interest_rate', 0)))
    cid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': cid}), 201

@app.route('/api/credit-cards/<int:cid>/payment', methods=['POST'])
@login_required
def create_cc_payment(cid):
    data = request.json or {}
    if not data.get('amount') or not data.get('date'):
        return jsonify({'error': 'Amount and date required'}), 400
    try:
        conn = get_db_connection()
        # Ownership check on the card
        can_see, can_edit = is_owner_or_shared(conn, 'credit_cards', cid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Credit card not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        conn.execute("INSERT INTO cc_payments (credit_card_id, amount, date, notes) VALUES (?,?,?,?)",
                  (cid, data['amount'], data['date'], data.get('notes')))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/credit-cards/<int:cid>', methods=['PUT'])
@login_required
def update_credit_card(cid):
    data = request.json or {}
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("UPDATE credit_cards SET owner=?,card_name=?,credit_limit=?,manual_balance=?,starting_balance=?,interest_rate=? WHERE id=?",
              (data.get('owner'), data.get('card_name'), data.get('credit_limit'), data.get('manual_balance'), data.get('starting_balance', 0), data.get('interest_rate', 0), cid))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200


@app.route('/api/credit-cards/<int:cid>', methods=['DELETE'])
@login_required
def delete_credit_card(cid):
    conn = get_db_connection()
    conn.execute("DELETE FROM credit_cards WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

# ── Budgets ───────────────────────────────────────────────────────────────────

@app.route('/api/budgets', methods=['GET'])
@login_required
def get_budgets():
    conn = get_db_connection()
    is_business = request.args.get('is_business', '0')
    vis_clause, vis_params = get_visible_clause('budgets', session['username'])
    q = f"SELECT * FROM budgets WHERE {vis_clause} AND is_business = ? ORDER BY year DESC, month DESC, category"
    rows = [dict(r) for r in conn.execute(q, vis_params + [int(is_business)]).fetchall()]
    conn.close()
    for b in rows:
        conn2 = get_db_connection()
        exp_vis, exp_params = get_visible_clause('expenses', session['username'])
        spent = conn2.execute(
            f"SELECT COALESCE(SUM(amount),0.0) FROM expenses WHERE category=? AND strftime('%Y',date)=? AND strftime('%m',date)=? AND is_business=? AND {exp_vis}",
            [b['category'], str(b['year']), f"{b['month']:02d}", int(is_business)] + exp_params).fetchone()[0]
        # Also get prev month spending for MoM comparison
        prev_month = b['month'] - 1 if b['month'] > 1 else 12
        prev_year = b['year'] if b['month'] > 1 else b['year'] - 1
        prev_spent = conn2.execute(
            f"SELECT COALESCE(SUM(amount),0.0) FROM expenses WHERE category=? AND strftime('%Y',date)=? AND strftime('%m',date)=? AND is_business=? AND {exp_vis}",
            [b['category'], str(prev_year), f"{prev_month:02d}", int(is_business)] + exp_params).fetchone()[0]
        conn2.close()
        effective = b['amount'] + b.get('rollover', 0)
        b['effective_amount'] = round(float(effective), 2)
        b['spent'] = round(float(spent), 2)
        b['prev_spent'] = round(float(prev_spent), 2)
        b['mom_change'] = round(((float(spent) - float(prev_spent)) / float(prev_spent) * 100), 1) if float(prev_spent) > 0 else None
        b['percentage'] = round(float(spent) / effective * 100, 1) if effective > 0 else 0
        b['is_over'] = b['percentage'] > 100
        b['rollover'] = round(float(b.get('rollover', 0)), 2)
    return jsonify({'data': rows}), 200

@app.route('/api/budgets', methods=['POST'])
@login_required
def create_budget():
    data = request.json or {}
    try:
        conn = get_db_connection()
        c = conn.cursor()
        # Always set owner to current user
        c.execute("INSERT OR REPLACE INTO budgets (category,amount,month,year,owner,is_business) VALUES (?,?,?,?,?,?)",
                  (data['category'], float(data['amount']), int(data['month']), int(data['year']), session['username'], int(data.get('is_business', 0))))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': c.lastrowid}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/budgets/<int:bid>', methods=['DELETE'])
@login_required
def delete_budget(bid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'budgets', bid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Budget not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        conn.execute("DELETE FROM budgets WHERE id=?", (bid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Business API ──────────────────────────────────────────────────────────────

@app.route('/api/business/stats')
@login_required
def get_business_stats():
    month = request.args.get('month')
    if not month:
        now = datetime.now()
        month = now.strftime('%Y-%m')
    
    conn = get_db_connection()
    username = session['username']
    vis_clause, vis_params = get_visible_clause('expenses', username)
    
    # Revenue (Income where is_business=1)
    revenue = conn.execute(
        f"SELECT COALESCE(SUM(amount), 0.0) FROM income WHERE is_business=1 AND date LIKE ? AND {vis_clause.replace('expenses', 'income')}",
        [month + '%'] + vis_params).fetchone()[0]
    
    # Expenses (Expenses where is_business=1)
    expenses = conn.execute(
        f"SELECT COALESCE(SUM(amount), 0.0) FROM expenses WHERE is_business=1 AND date LIKE ? AND {vis_clause}",
        [month + '%'] + vis_params).fetchone()[0]
    
    conn.close()
    
    profit = revenue - expenses
    margin = (profit / revenue * 100) if revenue > 0 else 0
    
    return jsonify({
        'data': {
            'revenue': round(revenue, 2),
            'expenses': round(expenses, 2),
            'profit': round(profit, 2),
            'margin': round(margin, 1)
        }
    }), 200

# ─── Business CSV Export ──────────────────────────────────────────────────────

@app.route('/api/business/export/csv')
@login_required
def export_business_csv():
    import csv, io
    username = session['username']
    conn = get_db_connection()
    vis_clause_exp, vis_params_exp = get_visible_clause('expenses', username)
    vis_clause_inc, vis_params_inc = get_visible_clause('income', username)

    expenses = conn.execute(
        f"SELECT date, category, description, amount, payment_method FROM expenses WHERE is_business=1 AND {vis_clause_exp} ORDER BY date DESC",
        vis_params_exp).fetchall()
    income = conn.execute(
        f"SELECT date, source, description, amount FROM income WHERE is_business=1 AND {vis_clause_inc} ORDER BY date DESC",
        vis_params_inc).fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['TYPE', 'DATE', 'CATEGORY/SOURCE', 'DESCRIPTION', 'AMOUNT', 'PAYMENT'])
    for e in expenses:
        writer.writerow(['Expense', e['date'], e['category'], e['description'], f"-{e['amount']:.2f}", e['payment_method'] or ''])
    for i in income:
        writer.writerow(['Income', i['date'], i['source'], i['description'], f"+{i['amount']:.2f}", ''])

    output.seek(0)
    from flask import make_response
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=business_report.csv'
    response.headers['Content-Type'] = 'text/csv'
    return response

# ─── Invoice API ──────────────────────────────────────────────────────────────

@app.route('/api/invoices', methods=['GET'])
@login_required
def get_invoices():
    username = session['username']
    try:
        conn = get_db_connection()
        # Strictly only owner can see invoices (no sharing for now)
        rows = conn.execute("SELECT * FROM invoices WHERE owner=? ORDER BY created_at DESC", (username,)).fetchall()
        conn.close()
        return jsonify({'data': [dict(r) for r in rows]}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/invoices', methods=['POST'])
@login_required
def create_invoice():
    import json as _json
    data = request.get_json()
    username = session['username']
    # Auto-generate invoice number
    conn = get_db_connection()
    count = conn.execute("SELECT COUNT(*) FROM invoices WHERE owner=?", (username,)).fetchone()[0]
    inv_number = f"INV-{datetime.now().strftime('%Y%m')}-{count+1:03d}"
    conn.execute("""INSERT INTO invoices
        (invoice_number, client_name, client_email, issue_date, due_date, items, notes, tax_rate, status, owner)
        VALUES (?,?,?,?,?,?,?,?,?,?)""", (
        inv_number,
        data.get('client_name', ''),
        data.get('client_email', ''),
        data.get('issue_date', ''),
        data.get('due_date', ''),
        _json.dumps(data.get('items', [])),
        data.get('notes', ''),
        float(data.get('tax_rate', 0)),
        'draft',
        username  # Strictly from session
    ))
    conn.commit()
    inv_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({'message': 'Invoice created', 'invoice_number': inv_number, 'id': inv_id}), 201

@app.route('/api/invoices/<int:inv_id>', methods=['PATCH'])
@login_required
def update_invoice_status(inv_id):
    data = request.get_json()
    status = data.get('status')
    if status not in ('draft', 'sent', 'paid'):
        return jsonify({'error': 'Invalid status'}), 400
    try:
        conn = get_db_connection()
        # Ownership check
        c = conn.cursor()
        row = c.execute("SELECT owner FROM invoices WHERE id=?", (inv_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Invoice not found'}), 404
        if row['owner'] != session['username']:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        conn.execute("UPDATE invoices SET status=? WHERE id=?", (status, inv_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Status updated'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/business/advisor')
@login_required
def get_business_advice():
    # Prompt Ollama for business advice
    location = request.args.get('location', 'Connecticut (CT)')
    ollama_url = Config.OLLAMA_URL
    if not ollama_url:
        return jsonify({'error': 'Ollama not configured'}), 400
    
    import requests
    
    username = session['username']
    conn = get_db_connection()
    month = datetime.now().strftime('%Y-%m')
    
    # Get some context
    revenue = conn.execute("SELECT COALESCE(SUM(amount),0) FROM income WHERE is_business=1 AND owner=?", (username,)).fetchone()[0]
    expenses = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE is_business=1 AND owner=?", (username,)).fetchone()[0]
    conn.close()
    
    system_prompt = (
        "You are Kash Business Advisor. Provide concise, expert advice on small business finances. "
        f"User is based in {location}. Focus on tax deadlines (e.g., quarterly estimated taxes or VAT), "
        "profitability, and business growth. If the location is in France, consider French tax laws like TVA and URSSAF. "
        "Format with markdown."
    )
    user_prompt = f"My business currently has ${revenue:,.2f} in revenue and ${expenses:,.2f} in expenses to date. It is currently {datetime.now().strftime('%B %Y')}. Give me some advice on what I should be looking out for regarding {location} business requirements or taxes."

    payload = {
        "model": Config.OLLAMA_MODEL,
        "prompt": f"{system_prompt}\n\nUser: {user_prompt}\n\nAdvisor:",
        "stream": False
    }
    
    try:
        r = requests.post(f"{ollama_url}/api/generate", json=payload, timeout=30)
        r.raise_for_status()
        advice = r.json().get('response', '')
        return jsonify({'advice': advice}), 200
    except Exception as e:
        return jsonify({'error': f"Failed to reach Ollama: {str(e)}"}), 500


@app.route('/api/budgets/copy-last-month', methods=['POST'])
@login_required
def copy_last_month_budgets():
    now = datetime.now()
    cur_month, cur_year = now.month, now.year
    prev_month = cur_month - 1 if cur_month > 1 else 12
    prev_year = cur_year if cur_month > 1 else cur_year - 1
    conn = get_db_connection()
    vis_clause, vis_params = get_visible_clause('budgets', session['username'])
    prev_budgets = conn.execute(
        f"SELECT category, amount, is_business FROM budgets WHERE month=? AND year=? AND {vis_clause}",
        [prev_month, prev_year] + vis_params).fetchall()
    if not prev_budgets:
        conn.close()
        return jsonify({'error': 'No budgets found for last month'}), 404
    copied = 0
    for b in prev_budgets:
        # Calculate rollover from prev month
        exp_vis, exp_params = get_visible_clause('expenses', session['username'])
        spent = conn.execute(
            f"SELECT COALESCE(SUM(amount),0.0) FROM expenses WHERE category=? AND strftime('%Y',date)=? AND strftime('%m',date)=? AND is_business=? AND {exp_vis}",
            [b['category'], str(prev_year), f"{prev_month:02d}", b['is_business']] + exp_params).fetchone()[0]
        rollover = max(0, b['amount'] - float(spent))
        try:
            conn.execute(
                "INSERT OR IGNORE INTO budgets (category,amount,month,year,rollover,owner,is_business) VALUES (?,?,?,?,?,?,?)",
                (b['category'], b['amount'], cur_month, cur_year, round(rollover, 2), session['username'], b['is_business']))
            copied += 1
        except Exception:
            pass
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'copied': copied}), 200

@app.route('/api/budgets/apply-rollover', methods=['POST'])
@login_required
def apply_rollover():
    """Manually apply rollover from previous month to current month budgets."""
    now = datetime.now()
    cur_month, cur_year = now.month, now.year
    prev_month = cur_month - 1 if cur_month > 1 else 12
    prev_year = cur_year if cur_month > 1 else cur_year - 1
    conn = get_db_connection()
    cur_budgets = conn.execute(
        "SELECT id, category, amount, is_business FROM budgets WHERE month=? AND year=?",
        (cur_month, cur_year)).fetchall()
    updated = 0
    for b in cur_budgets:
        prev = conn.execute(
            "SELECT amount FROM budgets WHERE category=? AND month=? AND year=? AND is_business=?",
            (b['category'], prev_month, prev_year, b['is_business'])).fetchone()
        if prev:
            prev_spent = conn.execute(
                "SELECT COALESCE(SUM(amount),0.0) FROM expenses WHERE category=? AND strftime('%Y',date)=? AND strftime('%m',date)=? AND is_business=?",
                (b['category'], str(prev_year), f"{prev_month:02d}", b['is_business'])).fetchone()[0]
            rollover = max(0, prev['amount'] - float(prev_spent))
            if rollover > 0:
                conn.execute("UPDATE budgets SET rollover=? WHERE id=?", (round(rollover, 2), b['id']))
                updated += 1
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'updated': updated}), 200

# ── Savings Goals ──────────────────────────────────────────────────────────────

@app.route('/api/savings-goals', methods=['GET'])
@login_required
def get_savings_goals():
    conn = get_db_connection()
    goals = [dict(r) for r in conn.execute("SELECT * FROM savings_goals ORDER BY target_date").fetchall()]
    # Calculate current savings (total income - total expenses)
    total_income = conn.execute("SELECT COALESCE(SUM(amount),0) FROM income WHERE is_business=0").fetchone()[0]
    total_expenses = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE is_business=0").fetchone()[0]
    current_savings = float(total_income) - float(total_expenses)
    conn.close()
    for g in goals:
        progress = min(100, round((current_savings / g['target_amount']) * 100, 1)) if g['target_amount'] > 0 else 0
        days_left = (datetime.strptime(g['target_date'], '%Y-%m-%d').date() - datetime.now().date()).days
        months_left = max(1, round(days_left / 30))
        remaining = max(0, g['target_amount'] - current_savings)
        g['current_savings'] = round(current_savings, 2)
        g['progress'] = progress
        g['days_left'] = days_left
        g['months_left'] = months_left
        g['monthly_needed'] = round(remaining / months_left, 2)
        g['on_track'] = current_savings >= g['target_amount']
    return jsonify({'data': goals, 'current_savings': round(current_savings, 2)}), 200

@app.route('/api/savings-goals', methods=['POST'])
@login_required
def create_savings_goal():
    data = request.json or {}
    if not data.get('name') or not data.get('target_amount') or not data.get('target_date'):
        return jsonify({'error': 'Name, target amount and date required'}), 400
    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("INSERT INTO savings_goals (name,target_amount,target_date,owner) VALUES (?,?,?,?)",
                  (data['name'], float(data['target_amount']), data['target_date'], session['username']))
        gid = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': gid}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/savings-goals/<int:gid>', methods=['DELETE'])
@login_required
def delete_savings_goal(gid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'savings_goals', gid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Savings goal not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        conn.execute("DELETE FROM savings_goals WHERE id=?", (gid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Loans ─────────────────────────────────────────────────────────────────────

@app.route('/api/loans', methods=['GET'])
@login_required
def get_loans():
    conn = get_db_connection()
    c = conn.cursor()
    # Basic visibility check
    vis_clause, vis_params = get_visible_clause('loans', session['username'])
    loans = [dict(r) for r in c.execute(f"SELECT * FROM loans WHERE {vis_clause} ORDER BY is_active DESC, created_at DESC", vis_params).fetchall()]
    
    for loan in loans:
        # Sum payments
        c.execute("SELECT COALESCE(SUM(amount),0.0) FROM loan_payments WHERE loan_id=?", (loan['id'],))
        paid = float(c.fetchone()[0])
        loan['amount_paid'] = round(paid, 2)
        loan['remaining_balance'] = round(loan['total_amount'] - paid, 2)
        loan['progress_pct'] = min(100, round((paid / loan['total_amount'] * 100), 1)) if loan['total_amount'] > 0 else 0
        
        # Get payment history
        payments = [dict(r) for r in c.execute("SELECT * FROM loan_payments WHERE loan_id=? ORDER BY date DESC", (loan['id'],)).fetchall()]
        loan['payments'] = payments

        # Amortization Calculation
        loan['payoff_months'] = None
        loan['payoff_date'] = None
        
        if loan['remaining_balance'] > 0 and loan['monthly_payment'] and loan['monthly_payment'] > 0:
            P = loan['remaining_balance']
            M = loan['monthly_payment']
            r = (loan.get('interest_rate', 0) or 0) / 100 / 12
            
            if r == 0:
                months = math.ceil(P / M)
                loan['payoff_months'] = months
            elif M > P * r:
                # Formula: n = -log(1 - (r*P)/M) / log(1+r)
                try:
                    months = -math.log(1 - (r * P) / M) / math.log(1 + r)
                    loan['payoff_months'] = math.ceil(months)
                except (ValueError, ZeroDivisionError):
                    pass
            
            if loan['payoff_months']:
                # Simple estimation: current month + months
                today = dt.date.today()
                payoff_dt = today + timedelta(days=loan['payoff_months'] * 30)
                loan['payoff_date'] = payoff_dt.strftime('%b %Y')

    conn.close()
    return jsonify({'data': loans}), 200

@app.route('/api/loans', methods=['POST'])
@login_required
def create_loan():
    data = request.json or {}
    if not data.get('loan_name') or not data.get('total_amount'):
        return jsonify({'error': 'Name and total amount required'}), 400
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""INSERT INTO loans (owner, loan_name, total_amount, monthly_payment, next_due_date, category, interest_rate, notes) 
                 VALUES (?,?,?,?,?,?,?,?)""",
              (session['username'], data['loan_name'], float(data['total_amount']), 
               data.get('monthly_payment'), data.get('next_due_date'), data.get('category'), data.get('interest_rate', 0), data.get('notes')))
    lid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': lid}), 201

@app.route('/api/loans/<int:lid>', methods=['PUT'])
@login_required
def update_loan(lid):
    data = request.json or {}
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'loans', lid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Loan not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        c = conn.cursor()
        c.execute("""UPDATE loans SET loan_name=?, total_amount=?, monthly_payment=?, 
                     next_due_date=?, category=?, interest_rate=?, notes=?, is_active=? WHERE id=?""",
                  (data.get('loan_name'), float(data['total_amount']), data.get('monthly_payment'),
                   data.get('next_due_date'), data.get('category'), data.get('interest_rate', 0), data.get('notes'), data.get('is_active', 1), lid))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/loans/<int:lid>', methods=['DELETE'])
@login_required
def delete_loan(lid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'loans', lid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Loan not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        conn.execute("DELETE FROM loans WHERE id=?", (lid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/loans/<int:lid>/payment', methods=['POST'])
@login_required
def create_loan_payment(lid):
    data = request.json or {}
    if not data.get('amount') or not data.get('date'):
        return jsonify({'error': 'Amount and date required'}), 400
    try:
        conn = get_db_connection()
        # Ownership check on the loan
        can_see, can_edit = is_owner_or_shared(conn, 'loans', lid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Loan not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        c = conn.cursor()
        # 1. Record the loan payment
        c.execute("INSERT INTO loan_payments (loan_id, amount, date, notes) VALUES (?,?,?,?)",
                  (lid, float(data['amount']), data['date'], data.get('notes')))
        
        # 2. Optionally record it as an expense if user wants
        if data.get('as_expense'):
            c.execute("SELECT loan_name FROM loans WHERE id=?", (lid,))
            loan_name = c.fetchone()[0]
            c.execute("INSERT INTO expenses (date,category,description,amount,owner) VALUES (?,?,?,?,?)",
                      (data['date'], 'Debt', f"Loan Payment: {loan_name}", float(data['amount']), session['username']))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        loan = c.execute("SELECT loan_name, category FROM loans WHERE id=?", (lid,)).fetchone()
        c.execute("""INSERT INTO expenses (date, category, description, amount, paid_by, notes) 
                     VALUES (?,?,?,?,?,?)""",
                  (data['date'], loan['category'] or 'Bills', f"Payment: {loan['loan_name']}", 
                   float(data['amount']), session['username'], data.get('notes')))
        
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 201


# ── Spending by Person ─────────────────────────────────────────────────────────

@app.route('/api/stats/spending-by-person')
@login_required
def spending_by_person():
    now = datetime.now()
    month = request.args.get('month', f"{now.year}-{now.month:02d}")
    year, mon = month.split('-')
    conn = get_db_connection()
    exp_vis, exp_params = get_visible_clause('expenses', session['username'])
    rows = conn.execute(
        f"""SELECT paid_by, COALESCE(SUM(amount),0) as total
           FROM expenses
           WHERE strftime('%Y',date)=? AND strftime('%m',date)=? AND is_business=? AND {exp_vis}
           GROUP BY paid_by ORDER BY total DESC""",
        [year, mon, int(request.args.get('is_business', '0'))] + exp_params).fetchall()
    conn.close()
    return jsonify({'data': [dict(r) for r in rows]}), 200

# ── Bills ─────────────────────────────────────────────────────────────────────

@app.route('/api/bills', methods=['GET'])
@login_required
def get_bills():
    is_business = request.args.get('is_business', '0')
    conn = get_db_connection()
    vis_clause, vis_params = get_visible_clause('bills', session['username'])
    q = f"SELECT * FROM bills WHERE {vis_clause} AND is_business = ? ORDER BY due_date ASC"
    rows = [dict(r) for r in conn.execute(q, vis_params + [int(is_business)]).fetchall()]
    conn.close()
    today = datetime.now().date()
    for b in rows:
        try:
            due = datetime.strptime(b['due_date'], '%Y-%m-%d').date()
            b['days_until'] = (due - today).days
        except:
            b['days_until'] = None
    return jsonify({'data': rows}), 200

@app.route('/api/bills', methods=['POST'])
@login_required
def create_bill():
    data = request.json or {}
    if not data.get('name') or not data.get('amount') or not data.get('due_date'):
        return jsonify({'error': 'Name, amount, and due_date required'}), 400
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("INSERT INTO bills (name,amount,due_date,category,notes,owner,is_business) VALUES (?,?,?,?,?,?,?)",
              (data['name'], float(data['amount']), data['due_date'], data.get('category'), data.get('notes',''), session['username'], data.get('is_business', 0)))
    bid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': bid}), 201

@app.route('/api/bills/<int:bid>', methods=['PUT'])
@login_required
def update_bill(bid):
    data = request.json or {}
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'bills', bid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Bill not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        c = conn.cursor()
        c.execute("UPDATE bills SET name=?,amount=?,due_date=?,category=?,is_paid=?,notes=?,is_business=? WHERE id=?",
                  (data.get('name'), data.get('amount'), data.get('due_date'),
                   data.get('category'), data.get('is_paid',0), data.get('notes',''), data.get('is_business', 0), bid))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bills/<int:bid>/toggle', methods=['POST'])
@login_required
def toggle_bill_paid(bid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'bills', bid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Bill not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        c = conn.cursor()
        c.execute("UPDATE bills SET is_paid = CASE WHEN is_paid=1 THEN 0 ELSE 1 END WHERE id=?", (bid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bills/<int:bid>', methods=['DELETE'])
@login_required
def delete_bill(bid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'bills', bid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Bill not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        conn.execute("DELETE FROM bills WHERE id=?", (bid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Recurring ─────────────────────────────────────────────────────────────────

@app.route('/api/recurring', methods=['GET'])
@login_required
def get_recurring():
    is_business = request.args.get('is_business', '0')
    conn = get_db_connection()
    vis_clause, vis_params = get_visible_clause('recurring_transactions', session['username'])
    q = f"SELECT * FROM recurring_transactions WHERE {vis_clause} AND is_business = ? ORDER BY next_date"
    rows = [dict(r) for r in conn.execute(q, vis_params + [int(is_business)]).fetchall()]
    conn.close()
    return jsonify({'data': rows}), 200

@app.route('/api/recurring', methods=['POST'])
@login_required
def create_recurring():
    data = request.json or {}
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""INSERT INTO recurring_transactions
                 (type,frequency,next_date,description,amount,category,source,person,payment_method,credit_card_id,owner,is_business)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
              (data.get('type'), data.get('frequency'), data.get('next_date'), data.get('description'),
               float(data.get('amount',0)), data.get('category'), data.get('source'),
               data.get('person'), data.get('payment_method'), data.get('credit_card_id'), session['username'], data.get('is_business', 0)))
    rid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': rid}), 201

@app.route('/api/recurring/<int:rid>/process', methods=['POST'])
@login_required
def process_recurring(rid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'recurring_transactions', rid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Recurring transaction not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        c = conn.cursor()
        c.execute("SELECT * FROM recurring_transactions WHERE id=?", (rid,))
        rec = c.fetchone()
        if not rec:
            conn.close()
            return jsonify({'error': 'Not found'}), 404
        rec = dict(rec)
        today = datetime.now().strftime('%Y-%m-%d')
        if rec['type'] == 'expense':
            c.execute("INSERT INTO expenses (date,category,description,amount,paid_by,payment_method,credit_card_id,notes,owner,is_business) VALUES (?,?,?,?,?,?,?,?,?,?)",
                      (today, rec['category'] or 'Other', rec['description'], rec['amount'],
                       rec['person'], rec['payment_method'], rec['credit_card_id'], 'Auto from recurring', rec['owner'], rec.get('is_business', 0)))
        else:
            c.execute("INSERT INTO income (date,source,description,amount,received_by,notes,owner,is_business) VALUES (?,?,?,?,?,?,?,?)",
                      (today, rec['source'] or 'Other', rec['description'], rec['amount'],
                       rec['person'], 'Auto from recurring', rec['owner'], rec.get('is_business', 0)))
        # Advance next_date
        freq = rec['frequency']
        from datetime import timedelta
        nd = datetime.strptime(rec['next_date'], '%Y-%m-%d')
        if freq == 'weekly': nd += timedelta(weeks=1)
        elif freq == 'biweekly': nd += timedelta(weeks=2)
        elif freq == 'monthly': nd = nd.replace(month=nd.month % 12 + 1) if nd.month < 12 else nd.replace(year=nd.year+1, month=1)
        elif freq == 'yearly': nd = nd.replace(year=nd.year+1)
        c.execute("UPDATE recurring_transactions SET next_date=? WHERE id=?", (nd.strftime('%Y-%m-%d'), rid))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'Processed and next date is {nd.strftime("%Y-%m-%d")}'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/recurring/<int:rid>', methods=['DELETE'])
@login_required
def delete_recurring(rid):
    try:
        conn = get_db_connection()
        # Ownership check
        can_see, can_edit = is_owner_or_shared(conn, 'recurring_transactions', rid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Recurring transaction not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        conn.execute("DELETE FROM recurring_transactions WHERE id=?", (rid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/expenses/<int:eid>/receipt', methods=['POST'])
@login_required
def upload_receipt(eid):
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    try:
        conn = get_db_connection()
        # Ownership check on the expense
        can_see, can_edit = is_owner_or_shared(conn, 'expenses', eid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Expense not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        import uuid
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'bin'
        filename = f"{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        c = conn.cursor()
        c.execute("UPDATE expenses SET receipt_filename=? WHERE id=?", (filename, eid))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'path': filename}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/expenses/<int:eid>/receipt', methods=['DELETE'])
@login_required
def delete_receipt(eid):
    try:
        conn = get_db_connection()
        # Ownership check on the expense
        can_see, can_edit = is_owner_or_shared(conn, 'expenses', eid, session['username'])
        if not can_see:
            conn.close()
            return jsonify({'error': 'Expense not found'}), 404
        if not can_edit:
            conn.close()
            return jsonify({'error': 'Permission denied'}), 403

        row = conn.execute("SELECT receipt_filename FROM expenses WHERE id=?", (eid,)).fetchone()
        if row and row['receipt_filename']:
            import os
            path = os.path.join(app.config['UPLOAD_FOLDER'], row['receipt_filename'])
            if os.path.exists(path):
                os.remove(path)
            conn.execute("UPDATE expenses SET receipt_filename=NULL WHERE id=?", (eid,))
            conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/receipts/<filename>')
@login_required
def serve_receipt(filename):
    try:
        conn = get_db_connection()
        # Find the expense associated with this receipt
        row = conn.execute("SELECT id FROM expenses WHERE receipt_filename=?", (filename,)).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Receipt not found'}), 404
        
        eid = row['id']
        # Ownership check
        can_see, _ = is_owner_or_shared(conn, 'expenses', eid, session['username'])
        conn.close()
        
        if not can_see:
            return jsonify({'error': 'Permission denied'}), 403
            
        import os
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Currencies ─────────────────────────────────────────────────────────────────

@app.route('/api/currencies', methods=['GET'])
@login_required
def get_currencies():
    conn = get_db_connection()
    rows = [dict(r) for r in conn.execute("SELECT * FROM currencies ORDER BY is_home DESC, code").fetchall()]
    conn.close()
    return jsonify({'data': rows}), 200

@app.route('/api/currencies', methods=['POST'])
@admin_required
def create_currency():
    data = request.json or {}
    conn = get_db_connection()
    try:
        conn.execute(
            "INSERT INTO currencies (code,name,symbol,rate_to_home,is_home) VALUES (?,?,?,?,?)",
            (data['code'].upper(), data['name'], data['symbol'],
             float(data.get('rate_to_home', 1.0)), 1 if data.get('is_home') else 0))
        if data.get('is_home'):
            conn.execute("UPDATE currencies SET is_home=0 WHERE code!=?", (data['code'].upper(),))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 201
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/currencies/<int:cid>', methods=['PUT'])
@admin_required
def update_currency(cid):
    data = request.json or {}
    conn = get_db_connection()
    conn.execute(
        "UPDATE currencies SET name=?,symbol=?,rate_to_home=?,updated_at=datetime('now') WHERE id=?",
        (data['name'], data['symbol'], float(data['rate_to_home']), cid))
    if data.get('is_home'):
        conn.execute("UPDATE currencies SET is_home=0")
        conn.execute("UPDATE currencies SET is_home=1 WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

@app.route('/api/currencies/<int:cid>', methods=['DELETE'])
@admin_required
def delete_currency(cid):
    conn = get_db_connection()
    home = conn.execute("SELECT is_home FROM currencies WHERE id=?", (cid,)).fetchone()
    if home and home['is_home']:
        conn.close()
        return jsonify({'error': 'Cannot delete home currency'}), 400
    conn.execute("DELETE FROM currencies WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

# ── Audit Log ──────────────────────────────────────────────────────────────────

@app.route('/api/audit-log')
@admin_required
def get_audit_log():
    limit = int(request.args.get('limit', 100))
    table = request.args.get('table', '')
    conn = get_db_connection()
    q = "SELECT * FROM audit_log WHERE 1=1"
    params = []
    if table:
        q += " AND table_name=?"; params.append(table)
    q += f" ORDER BY created_at DESC LIMIT {limit}"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()
    return jsonify({'data': rows}), 200

@app.route('/api/expenses/check-duplicate', methods=['POST'])
@login_required
def check_duplicate():
    data = request.json or {}
    from datetime import timedelta as _td
    try:
        from_date = datetime.strptime(data['date'], '%Y-%m-%d')
    except Exception:
        return jsonify({'duplicate': False}), 200
    d1 = (from_date - _td(days=3)).strftime('%Y-%m-%d')
    d2 = (from_date + _td(days=3)).strftime('%Y-%m-%d')
    conn = get_db_connection()
    dup = conn.execute(
        "SELECT id,date,description,amount FROM expenses WHERE description=? AND amount=? AND date BETWEEN ? AND ?",
        (data.get('description',''), float(data.get('amount',0)), d1, d2)).fetchone()
    conn.close()
    if dup:
        return jsonify({'duplicate': True, 'existing': dict(dup)}), 200
    return jsonify({'duplicate': False}), 200

# ── Dashboard with month param ─────────────────────────────────────────────────

@app.route('/api/stats/summary-by-month')
@login_required
def get_summary_by_month():
    now = datetime.now()
    month = request.args.get('month', f"{now.year}-{now.month:02d}")
    conn = get_db_connection()
    def qval(sql, params=()):
        return float(conn.execute(sql, params).fetchone()[0])
    username = session['username']
    exp_vis, exp_params = get_visible_clause('expenses', username)
    inc_vis, inc_params = get_visible_clause('income', username)
    bill_vis, bill_params = get_visible_clause('bills', username)
    monthly_expenses = qval(f"SELECT COALESCE(SUM(amount),0) FROM expenses WHERE strftime('%Y-%m',date)=? AND is_business=0 AND {exp_vis}", [month] + exp_params)
    monthly_income   = qval(f"SELECT COALESCE(SUM(amount),0) FROM income WHERE strftime('%Y-%m',date)=? AND is_business=0 AND {inc_vis}", [month] + inc_params)
    total_income     = qval(f"SELECT COALESCE(SUM(amount),0) FROM income WHERE is_business=0 AND {inc_vis}", inc_params)
    unpaid_bills     = qval(f"SELECT COALESCE(SUM(amount),0) FROM bills WHERE is_paid=0 AND is_business=0 AND {bill_vis}", bill_params)
    cc_vis, cc_params = get_visible_clause('credit_cards', username)
    card_ids         = [r[0] for r in conn.execute(f"SELECT id FROM credit_cards WHERE {cc_vis}", cc_params).fetchall()]
    total_cc = sum(get_credit_card_balance(cid) for cid in card_ids)

    # Debt Payoff Projections
    loan_vis, loan_params = get_visible_clause('loans', username)
    all_loans = conn.execute(f"SELECT total_amount, monthly_payment, interest_rate, id FROM loans WHERE is_active=1 AND {loan_vis}", loan_params).fetchall()
    total_debt = total_cc
    max_months = 0
    
    for L in all_loans:
        paid = conn.execute("SELECT COALESCE(SUM(amount),0.0) FROM loan_payments WHERE loan_id=?", (L['id'],)).fetchone()[0]
        rem = L['total_amount'] - paid
        total_debt += rem
        if L['monthly_payment'] and L['monthly_payment'] > 0:
            P, M, r = rem, L['monthly_payment'], (L['interest_rate'] or 0)/100/12
            if r == 0: m = math.ceil(P/M)
            elif M > P*r:
                try: m = math.ceil(-math.log(1-(r*P)/M)/math.log(1+r))
                except: m = 0
            else: m = 120 # infinity
            if m > max_months: max_months = m

    conn.close()
    debt_free_date = "N/A"
    if max_months > 0:
        target = datetime.now() + timedelta(days=max_months*30)
        debt_free_date = target.strftime('%b %Y')

    return jsonify({
        'monthly_expenses': round(monthly_expenses, 2),
        'monthly_income': round(monthly_income, 2),
        'total_income': round(total_income, 2),
        'total_credit_card_balance': round(total_cc, 2),
        'total_debt': round(total_debt, 2),
        'debt_free_date': debt_free_date,
        'unpaid_bills_total': round(unpaid_bills, 2),
        'left_after_bills': round(monthly_income - unpaid_bills, 2)
    }), 200


# ── Spending Insights ─────────────────────────────────────────────────────────

@app.route('/api/insights')
@login_required
def get_insights():
    is_business = request.args.get('is_business', '0')
    conn = get_db_connection()
    now = datetime.now()
    cur_month = f"{now.year}-{now.month:02d}"
    # Prev 3 months for averages
    months = []
    y, m = now.year, now.month
    for _ in range(3):
        m -= 1
        if m == 0: m = 12; y -= 1
        months.append(f"{y}-{m:02d}")

    insights = []

    # Days elapsed this month
    days_elapsed = now.day
    days_in_month = 30  # approximate

    exp_vis, exp_params = get_visible_clause('expenses', session['username'])
    # Current month spending
    cur_spend = float(conn.execute(
        f"SELECT COALESCE(SUM(amount),0) FROM expenses WHERE strftime('%Y-%m',date)=? AND is_business=? AND {exp_vis}",
        [cur_month, int(is_business)] + exp_params).fetchone()[0])

    # Forecast: linear projection
    if days_elapsed > 0:
        daily_rate = cur_spend / days_elapsed
        forecast = round(daily_rate * days_in_month, 2)
        insights.append({
            'type': 'forecast',
            'icon': '📈',
            'title': 'Month-End Forecast',
            'message': f"At your current pace you'll spend ${forecast:,.2f} this month.",
            'value': forecast,
            'severity': 'info'
        })

    # Category anomalies — compare current month vs 3-month average
    cur_cats = {r[0]: float(r[1]) for r in conn.execute(
        f"SELECT category, SUM(amount) FROM expenses WHERE strftime('%Y-%m',date)=? AND is_business=? AND {exp_vis} GROUP BY category",
        [cur_month, int(is_business)] + exp_params).fetchall()}

    for cat, cur_amt in cur_cats.items():
        if not months: continue
        hist_rows = conn.execute(
            f"SELECT COALESCE(AVG(monthly),0) FROM (SELECT SUM(amount) as monthly FROM expenses WHERE category=? AND strftime('%Y-%m',date) IN ({','.join('?'*len(months))}) AND is_business=? AND {exp_vis} GROUP BY strftime('%Y-%m',date))",
            [cat, *months, int(is_business)] + exp_params).fetchone()
        avg = float(hist_rows[0]) if hist_rows and hist_rows[0] else 0
        if avg > 10 and cur_amt > avg * 1.4:
            pct = round(((cur_amt - avg) / avg) * 100)
            insights.append({
                'type': 'anomaly',
                'icon': '⚠️',
                'title': f'{cat} Spending Up',
                'message': f"You've spent ${cur_amt:,.2f} on {cat} this month — {pct}% more than your usual ${avg:,.2f}.",
                'severity': 'warning'
            })
        elif avg > 10 and cur_amt < avg * 0.6:
            pct = round(((avg - cur_amt) / avg) * 100)
            insights.append({
                'type': 'saving',
                'icon': '✅',
                'title': f'Saving on {cat}',
                'message': f"Great job! You're {pct}% under your usual {cat} spending.",
                'severity': 'success'
            })

    # Largest single expense this month
    biggest = conn.execute(
        f"SELECT description, amount FROM expenses WHERE strftime('%Y-%m',date)=? AND {exp_vis} ORDER BY amount DESC LIMIT 1",
        [cur_month] + exp_params).fetchone()
    if biggest and float(biggest[1]) > 50:
        insights.append({
            'type': 'largest',
            'icon': '💸',
            'title': 'Biggest Expense',
            'message': f'Your largest expense this month was "{biggest[0]}" at ${biggest[1]:,.2f}.',
            'severity': 'info'
        })

    # Unpaid overdue bills
    bill_vis, bill_params = get_visible_clause('bills', session['username'])
    overdue = conn.execute(
        f"SELECT COUNT(*) FROM bills WHERE is_paid=0 AND due_date < date('now') AND {bill_vis}", bill_params
    ).fetchone()[0]
    if overdue > 0:
        insights.append({
            'type': 'overdue',
            'icon': '🚨',
            'title': 'Overdue Bills',
            'message': f'You have {overdue} overdue bill{"s" if overdue>1 else ""}. Pay them to avoid late fees.',
            'severity': 'danger'
        })

    # Streak: consecutive months where saved > 0
    streak = 0
    sy, sm = now.year, now.month
    for _ in range(12):
        sm -= 1
        if sm == 0: sm = 12; sy -= 1
        mn = f"{sy}-{sm:02d}"
        inc_vis, inc_params = get_visible_clause('income', session['username'])
        inc = float(conn.execute(f"SELECT COALESCE(SUM(amount),0) FROM income WHERE strftime('%Y-%m',date)=? AND is_business=? AND {inc_vis}", [mn, int(is_business)] + inc_params).fetchone()[0])
        exp = float(conn.execute(f"SELECT COALESCE(SUM(amount),0) FROM expenses WHERE strftime('%Y-%m',date)=? AND is_business=? AND {exp_vis}", [mn, int(is_business)] + exp_params).fetchone()[0])
        if inc - exp > 0:
            streak += 1
        else:
            break
    if streak >= 2:
        insights.append({
            'type': 'streak',
            'icon': '🔥',
            'title': f'{streak}-Month Savings Streak',
            'message': f"You've saved money for {streak} months in a row. Keep it up!",
            'severity': 'success'
        })

    conn.close()
    forecast_val = forecast if days_elapsed > 0 else None
    return jsonify({'data': insights, 'forecast': forecast_val}), 200


# ── Category Sparklines ───────────────────────────────────────────────────────

@app.route('/api/stats/category-sparklines')
@login_required
def category_sparklines():
    is_business = request.args.get('is_business', '0')
    conn = get_db_connection()
    now = datetime.now()
    # Last 6 months
    months = []
    y, m = now.year, now.month
    for _ in range(6):
        months.append(f"{y}-{m:02d}")
        m -= 1
        if m == 0: m = 12; y -= 1
    months.reverse()

    exp_vis, exp_params = get_visible_clause('expenses', session['username'])
    categories = [r[0] for r in conn.execute(f"SELECT DISTINCT category FROM expenses WHERE {exp_vis} ORDER BY category", exp_params).fetchall()]
    result = {}
    for cat in categories:
        vals = []
        for mn in months:
            v = conn.execute(
                f"SELECT COALESCE(SUM(amount),0) FROM expenses WHERE category=? AND strftime('%Y-%m',date)=? AND is_business=? AND {exp_vis}",
                [cat, mn, int(is_business)] + exp_params).fetchone()[0]
            vals.append(round(float(v), 2))
        result[cat] = {'months': months, 'values': vals}
    conn.close()
    return jsonify({'data': result}), 200


# ── Bill Splits ───────────────────────────────────────────────────────────────

@app.route('/api/expenses/<int:eid>/splits', methods=['GET'])
@login_required
def get_splits(eid):
    conn = get_db_connection()
    splits = [dict(r) for r in conn.execute(
        "SELECT * FROM expense_splits WHERE expense_id=? ORDER BY username", (eid,)).fetchall()]
    conn.close()
    return jsonify({'data': splits}), 200

@app.route('/api/expenses/<int:eid>/splits', methods=['POST'])
@login_required
def save_splits(eid):
    data = request.json or {}
    splits = data.get('splits', [])
    conn = get_db_connection()
    conn.execute("DELETE FROM expense_splits WHERE expense_id=?", (eid,))
    for s in splits:
        if float(s.get('amount', 0)) > 0:
            conn.execute(
                "INSERT INTO expense_splits (expense_id,username,split_amount,split_pct) VALUES (?,?,?,?)",
                (eid, s['username'], float(s['amount']), s.get('pct')))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

@app.route('/api/expenses/split/<int:sid>/settle', methods=['POST'])
@login_required
def settle_split(sid):
    conn = get_db_connection()
    conn.execute("UPDATE expense_splits SET is_settled=1 WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

@app.route('/api/splits/summary')
@login_required
def splits_summary():
    """Who owes what to whom across all unsettled splits."""
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT s.username, s.split_amount, e.paid_by, e.description, s.id, s.expense_id
        FROM expense_splits s
        JOIN expenses e ON e.id = s.expense_id
        WHERE s.is_settled = 0 AND s.username != e.paid_by
        ORDER BY s.username
    """).fetchall()
    conn.close()
    return jsonify({'data': [dict(r) for r in rows]}), 200


# ── Widget Preferences ────────────────────────────────────────────────────────

@app.route('/api/widgets', methods=['GET'])
@login_required
def get_widget_prefs():
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT widget_id, is_visible, sort_order FROM widget_prefs WHERE username=? ORDER BY sort_order",
        (session['username'],)).fetchall()
    conn.close()
    return jsonify({'data': [dict(r) for r in rows]}), 200

@app.route('/api/widgets', methods=['POST'])
@login_required
def save_widget_prefs():
    data = request.json or {}
    widgets = data.get('widgets', [])
    conn = get_db_connection()
    for i, w in enumerate(widgets):
        conn.execute("""INSERT INTO widget_prefs (username,widget_id,is_visible,sort_order)
                        VALUES (?,?,?,?)
                        ON CONFLICT(username,widget_id) DO UPDATE SET is_visible=?,sort_order=?""",
                     (session['username'], w['id'], 1 if w.get('visible') else 0, i,
                      1 if w.get('visible') else 0, i))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200


# ── Onboarding ────────────────────────────────────────────────────────────────

@app.route('/api/onboarding', methods=['GET'])
@login_required
def get_onboarding():
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM onboarding WHERE username=?", (session['username'],)).fetchone()
    conn.close()
    if row:
        return jsonify({'completed': bool(row['completed']), 'step': row['step']}), 200
    return jsonify({'completed': False, 'step': 0}), 200

@app.route('/api/onboarding', methods=['POST'])
@login_required
def save_onboarding():
    data = request.json or {}
    conn = get_db_connection()
    completed = data.get('completed', False)
    step = data.get('step', 0)
    conn.execute("""INSERT INTO onboarding (username, completed, step, completed_at)
                    VALUES (?,?,?,?)
                    ON CONFLICT(username) DO UPDATE SET completed=?, step=?, completed_at=?""",
                 (session['username'], 1 if completed else 0, step,
                  datetime.now().isoformat() if completed else None,
                  1 if completed else 0, step,
                  datetime.now().isoformat() if completed else None))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

# ── Advanced Search ───────────────────────────────────────────────────────────

@app.route('/api/search/advanced')
@login_required
def advanced_search():
    q      = request.args.get('q', '').strip()
    cat    = request.args.get('category', '')
    person = request.args.get('person', '')
    method = request.args.get('method', '')
    start  = request.args.get('start', '')
    end    = request.args.get('end', '')
    amt_min= request.args.get('amt_min', '')
    amt_max= request.args.get('amt_max', '')
    limit  = int(request.args.get('limit', 100))

    is_business = int(request.args.get('is_business', 0))
    sql = """SELECT id,date,category,description,amount,paid_by,payment_method,
                    receipt_filename,notes,currency,original_amount
             FROM expenses WHERE is_business=? AND 1=1"""
    params = [is_business]
    if q:      sql += " AND (description LIKE ? OR notes LIKE ?)"; params += [f'%{q}%', f'%{q}%']
    if cat:    sql += " AND category=?"; params.append(cat)
    if person: sql += " AND paid_by=?"; params.append(person)
    if method: sql += " AND payment_method=?"; params.append(method)
    if start:  sql += " AND date>=?"; params.append(start)
    if end:    sql += " AND date<=?"; params.append(end)
    if amt_min:sql += " AND amount>=?"; params.append(float(amt_min))
    if amt_max:sql += " AND amount<=?"; params.append(float(amt_max))
    sql += f" ORDER BY date DESC LIMIT {limit}"

    conn = get_db_connection()
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    sum_sql = sql.replace("SELECT id,date,category,description,amount,paid_by,payment_method,receipt_filename,notes,currency,original_amount", "SELECT COALESCE(SUM(amount),0)").split("ORDER BY")[0]
    total = float(conn.execute(sum_sql, params).fetchone()[0])
    conn.close()
    return jsonify({'data': rows, 'count': len(rows), 'total': round(total, 2)}), 200


# ── Sharing ───────────────────────────────────────────────────────────────────

@app.route('/api/sharing/<table_name>/<int:record_id>', methods=['GET'])
@login_required
def get_sharing(table_name, record_id):
    """Get who a record is shared with."""
    conn = get_db_connection()
    can_see, _ = is_owner_or_shared(conn, table_name, record_id, session['username'])
    if not can_see:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    shares = [dict(r) for r in conn.execute(
        "SELECT * FROM sharing WHERE table_name=? AND record_id=?",
        (table_name, record_id)).fetchall()]
    # Get owner
    row = conn.execute(f"SELECT owner FROM {table_name} WHERE id=?", (record_id,)).fetchone()
    owner = row['owner'] if row else ''
    conn.close()
    return jsonify({'data': shares, 'owner': owner}), 200

@app.route('/api/sharing/<table_name>/<int:record_id>', methods=['POST'])
@login_required
def share_record(table_name, record_id):
    """Share a record with a user."""
    allowed_tables = ['expenses','income','credit_cards','bills','budgets','savings_goals']
    if table_name not in allowed_tables:
        return jsonify({'error': 'Invalid table'}), 400
    data = request.json or {}
    share_with = data.get('share_with', '').strip()
    can_edit   = 1 if data.get('can_edit', True) else 0
    if not share_with:
        return jsonify({'error': 'share_with is required'}), 400
    conn = get_db_connection()
    # Only owner can share
    row = conn.execute(f"SELECT owner FROM {table_name} WHERE id=?", (record_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Record not found'}), 404
    if row['owner'] != session['username'] and row['owner'] != '':
        conn.close()
        return jsonify({'error': 'Only the owner can share this record'}), 403
    # Verify user exists
    user_exists = conn.execute("SELECT id FROM users WHERE username=?", (share_with,)).fetchone()
    if not user_exists:
        conn.close()
        return jsonify({'error': f'User "{share_with}" not found'}), 404
    conn.execute("""INSERT INTO sharing (table_name,record_id,owner,shared_with,can_edit)
                    VALUES (?,?,?,?,?)
                    ON CONFLICT(table_name,record_id,shared_with) DO UPDATE SET can_edit=?""",
                 (table_name, record_id, session['username'], share_with, can_edit, can_edit))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 201

@app.route('/api/sharing/<table_name>/<int:record_id>/<shared_with>', methods=['DELETE'])
@login_required
def unshare_record(table_name, record_id, shared_with):
    """Remove sharing for a specific user."""
    conn = get_db_connection()
    row = conn.execute(f"SELECT owner FROM {table_name} WHERE id=?", (record_id,)).fetchone()
    if not row or (row['owner'] != session['username'] and row['owner'] != ''):
        conn.close()
        return jsonify({'error': 'Only the owner can unshare this record'}), 403
    conn.execute("DELETE FROM sharing WHERE table_name=? AND record_id=? AND shared_with=?",
                 (table_name, record_id, shared_with))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

@app.route('/api/sharing/summary')
@login_required
def sharing_summary():
    """Get all records shared with me and all records I have shared."""
    username = session['username']
    conn = get_db_connection()
    shared_with_me = [dict(r) for r in conn.execute(
        "SELECT * FROM sharing WHERE shared_with=? ORDER BY created_at DESC", (username,)).fetchall()]
    shared_by_me = [dict(r) for r in conn.execute(
        "SELECT * FROM sharing WHERE owner=? ORDER BY created_at DESC", (username,)).fetchall()]
    conn.close()
    return jsonify({'shared_with_me': shared_with_me, 'shared_by_me': shared_by_me}), 200


# ── Import ────────────────────────────────────────────────────────────────────

@app.route('/api/import/confirm', methods=['POST'])
@login_required
def import_confirm():
    """Save selected transactions, skipping duplicates."""
    data = request.json or {}
    rows = data.get('rows', [])
    paid_by = session['username']

    conn = get_db_connection()
    imported = 0
    skipped  = 0

    for row in rows:
        if not row.get('import'):
            continue
        tx_type  = row.get('type', 'expense')
        date     = row.get('date', '')
        desc     = row.get('description', '')
        amount   = float(row.get('amount', 0))
        category = row.get('category', 'Other')

        if not date or not desc or not amount:
            continue

        if tx_type == 'expense':
            # Check duplicate: same date + description + amount
            exists = conn.execute(
                "SELECT id FROM expenses WHERE date=? AND description=? AND amount=?",
                (date, desc, amount)).fetchone()
            if exists:
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO expenses (date,category,description,amount,paid_by,owner) VALUES (?,?,?,?,?,?)",
                (date, category, desc, amount, paid_by, paid_by))
        else:
            exists = conn.execute(
                "SELECT id FROM income WHERE date=? AND description=? AND amount=?",
                (date, desc, amount)).fetchone()
            if exists:
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO income (date,source,description,amount,received_by,owner) VALUES (?,?,?,?,?,?)",
                (date, category, desc, amount, paid_by, paid_by))
        imported += 1

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'imported': imported, 'skipped': skipped}), 200


@app.route('/api/import/ollama-categorize', methods=['POST'])
@login_required
def ollama_categorize():
    """Send descriptions to Ollama for AI categorization."""
    ollama_url = Config.OLLAMA_URL
    if not ollama_url:
        return jsonify({'error': 'Ollama not configured'}), 400

    data = request.json or {}
    descriptions = data.get('descriptions', [])
    categories   = data.get('categories', [])

    if not descriptions:
        return jsonify({'error': 'No descriptions provided'}), 400

    import urllib.request, json as jsonlib
    prompt = f"""You are a financial transaction categorizer. 
Given these categories: {', '.join(categories)}
Categorize each transaction description below. Reply with ONLY a JSON array of category strings, one per transaction, in the same order.
Descriptions:
{chr(10).join(f'{i+1}. {d}' for i, d in enumerate(descriptions))}
Reply with ONLY a JSON array like: ["Groceries","Dining Out","Transportation"]"""

    try:
        payload = jsonlib.dumps({
            'model': Config.OLLAMA_MODEL,
            'prompt': prompt,
            'stream': False
        }).encode('utf-8')
        url = ollama_url.rstrip('/')
        if not url.endswith('/api/generate'):
            url = f"{url}/api/generate"
        req = urllib.request.Request(
            url,
            data=payload,
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = jsonlib.loads(resp.read())
        response_text = result.get('response', '[]').strip()
        # Extract JSON array from response
        import re
        match = re.search(r'\[.*?\]', response_text, re.DOTALL)
        if match:
            cats = jsonlib.loads(match.group())
            return jsonify({'categories': cats}), 200
        return jsonify({'error': 'Could not parse Ollama response'}), 500
    except Exception as e:
        return jsonify({'error': f'Ollama request failed: {str(e)}'}), 500


@app.route('/api/import/ollama-status')
@login_required
def ollama_status():
    """Check if Ollama is configured and reachable."""
    ollama_url = Config.OLLAMA_URL
    if not ollama_url:
        return jsonify({'configured': False, 'url': ''}), 200
    import urllib.request
    try:
        url = ollama_url.rstrip('/')
        if url.endswith('/api/generate'):
            url = url[:-13]
        with urllib.request.urlopen(f"{url}/api/tags", timeout=5) as resp:
            return jsonify({'configured': True, 'url': ollama_url, 'reachable': True}), 200
    except:
        return jsonify({'configured': True, 'url': ollama_url, 'reachable': False}), 200

# ── AI Advisor ────────────────────────────────────────────────────────────────
@app.route('/api/advisor/plan', methods=['GET'])
@login_required
def advisor_plan():
    """Generates an AI financial plan based on the user's data."""
    ollama_url = Config.OLLAMA_URL
    if not ollama_url:
        return jsonify({'error': 'Ollama is not configured.'}), 400

    try:
        conn = get_db_connection()
        now = datetime.now()
        month = now.strftime('%Y-%m')
        username = session['username']
    
        # Gather data
        income_rows = conn.execute("SELECT SUM(amount) as total FROM income WHERE strftime('%Y-%m', date) = ? AND is_business=0", (month,)).fetchone()
        income = income_rows['total'] or 0.0
    
        # Gather fixed bills
        bills = conn.execute("SELECT name, amount FROM bills WHERE is_business=0").fetchall()
        bill_str = ", ".join([f"{b['name']} (${b['amount']})" for b in bills]) or "None"

        # Gather accurate Credit Card data
        vis_c, params_c = get_visible_clause('credit_cards', username)
        cc_rows = conn.execute(f"SELECT id, card_name, interest_rate FROM credit_cards WHERE {vis_c}", params_c).fetchall()
        cc_list = []
        for cr in cc_rows:
            bal = get_credit_card_balance(cr['id'])
            if bal > 0:
                cc_list.append(f"{cr['card_name']} (Balance: ${bal:,.2f}, {cr['interest_rate']}% APY)")
        debt_str = ", ".join(cc_list) or "None"

        # Gather Loan data
        vis_l, params_l = get_visible_clause('loans', username)
        loan_rows = conn.execute(f"SELECT id, loan_name, total_amount, monthly_payment, interest_rate FROM loans WHERE {vis_l} AND is_active=1", params_l).fetchall()
        loan_list = []
        for lr in loan_rows:
            # Calculate remaining balance
            paid = conn.execute("SELECT COALESCE(SUM(amount),0.0) FROM loan_payments WHERE loan_id=?", (lr['id'],)).fetchone()[0]
            rem = lr['total_amount'] - paid
            if rem > 0:
                loan_list.append(f"{lr['loan_name']} (Remaining: ${rem:,.2f}, {lr['interest_rate']}% APY, Min: ${lr['monthly_payment'] or 0}/mo)")
        loan_str = ", ".join(loan_list) or "None"
        
        # Gather Budgets
        budgets = conn.execute("SELECT category, amount FROM budgets WHERE year = ? AND month = ? AND is_business=0", (now.year, now.month)).fetchall()
        budget_str = ", ".join([f"{bg['category']} (${bg['amount']})" for bg in budgets]) or "None"
        
        conn.close()
    
        import urllib.request, urllib.error, json as jsonlib
        prompt = f"""You are an expert financial advisor. Provide a concise, actionable financial plan in markdown format.

User's Data for {month}:
- Total Monthly Income: ${income:,.2f}
- Monthly Fixed Bills: {bill_str}
- Credit Card Debt: {debt_str}
- Installment Loans: {loan_str}
- Current Budget Limits: {budget_str}

Please advise on:
1. Debt Payoff Strategy: Compare Debt Snowball, Debt Avalanche, and Velocity Banking for this user. Which is most efficient mathematically vs. psychologically? Provide a prioritized payoff order.
2. Budget Adjustments: Highlight specific areas where they can "find" money to accelerate the plan.
3. Savings & Cash Flow: How much can they realistically save? Should they build a starter emergency fund before attacking debt?

Your response MUST be exclusively formatted in nice Markdown with clear headings. Do not include introductory text like "Sure, here is your plan" - start immediately with the markdown content."""

        payload = jsonlib.dumps({
            'model': Config.OLLAMA_MODEL,
            'prompt': prompt,
            'stream': False
        }).encode('utf-8')
        url = ollama_url.rstrip('/')
        if not url.endswith('/api/generate'):
            url = f"{url}/api/generate"
        req = urllib.request.Request(
            url,
            data=payload,
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=90) as resp:
            result = jsonlib.loads(resp.read())
        response_text = result.get('response', '').strip()
        
        if not response_text:
            return jsonify({'error': 'Received an empty response from the AI.'}), 500
        return jsonify({'plan': response_text}), 200
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode('utf-8')
            return jsonify({'error': f'Ollama HTTP Error {e.code}: {err_body}'}), 500
        except:
            return jsonify({'error': f'Ollama HTTP Error {e.code}: {e.reason}'}), 500
    except Exception as e:
        return jsonify({'error': f'Failed to generate plan: {str(e)}'}), 500

# ── Stats ─────────────────────────────────────────────────────────────────────

@app.route('/api/stats/summary')
@login_required
def get_summary_stats():
    is_business = request.args.get('is_business', '0')
    conn = get_db_connection()
    now = datetime.now()
    current_month = f"{now.year}-{now.month:02d}"
    username = session['username']
    exp_vis, exp_params = get_visible_clause('expenses', username)
    inc_vis, inc_params = get_visible_clause('income', username)
    bill_vis, bill_params = get_visible_clause('bills', username)

    def qval(sql, params=()):
        return float(conn.execute(sql, params).fetchone()[0])

    total_expenses   = qval(f"SELECT COALESCE(SUM(amount),0) FROM expenses WHERE is_business=? AND {exp_vis}", [int(is_business)] + exp_params)
    total_income     = qval(f"SELECT COALESCE(SUM(amount),0) FROM income WHERE is_business=? AND {inc_vis}", [int(is_business)] + inc_params)
    monthly_expenses = qval(f"SELECT COALESCE(SUM(amount),0) FROM expenses WHERE strftime('%Y-%m',date)=? AND is_business=? AND {exp_vis}", (current_month, int(is_business)) + tuple(exp_params))
    monthly_income   = qval(f"SELECT COALESCE(SUM(amount),0) FROM income WHERE strftime('%Y-%m',date)=? AND is_business=? AND {inc_vis}", (current_month, int(is_business)) + tuple(inc_params))
    unpaid_bills     = qval(f"SELECT COALESCE(SUM(amount),0) FROM bills WHERE is_paid=0 AND is_business=? AND {bill_vis}", [int(is_business)] + bill_params)
    
    cc_vis, cc_params = get_visible_clause('credit_cards', username)
    card_ids         = [r[0] for r in conn.execute(f"SELECT id FROM credit_cards WHERE {cc_vis}", cc_params).fetchall()]
    
    total_cc = sum(get_credit_card_balance(cid) for cid in card_ids)
    
    # Total Debt Payoff Projection (Simpler approximation)
    loan_vis, loan_params = get_visible_clause('loans', username)
    all_loans = conn.execute(f"SELECT total_amount, monthly_payment, interest_rate, id FROM loans WHERE is_active=1 AND {loan_vis}", loan_params).fetchall()
    total_debt = total_cc
    total_monthly_pmt = 0
    max_months = 0
    
    for L in all_loans:
        paid = conn.execute("SELECT COALESCE(SUM(amount),0.0) FROM loan_payments WHERE loan_id=?", (L['id'],)).fetchone()[0]
        rem = L['total_amount'] - paid
        total_debt += rem
        if L['monthly_payment'] and L['monthly_payment'] > 0:
            total_monthly_pmt += L['monthly_payment']
            # Individual loan payoff for max_months
            P, M, r = rem, L['monthly_payment'], (L['interest_rate'] or 0)/100/12
            if r == 0: m = math.ceil(P/M)
            elif M > P*r:
                try: m = math.ceil(-math.log(1-(r*P)/M)/math.log(1+r))
                except: m = 0
            else: m = 120 # infinity/long
            if m > max_months: max_months = m

    conn.close()
    
    debt_free_date = "N/A"
    if max_months > 0:
        from datetime import timedelta
        target = datetime.now() + timedelta(days=max_months*30)
        debt_free_date = target.strftime('%b %Y')

    return jsonify({
        'total_expenses': round(total_expenses, 2),
        'total_income': round(total_income, 2),
        'monthly_expenses': round(monthly_expenses, 2),
        'monthly_income': round(monthly_income, 2),
        'total_credit_card_balance': round(total_cc, 2),
        'total_debt': round(total_debt, 2),
        'debt_free_date': debt_free_date,
        'unpaid_bills_total': round(unpaid_bills, 2),
        'left_after_bills': round(monthly_income - unpaid_bills, 2)
    }), 200

@app.route('/api/stats/spending-by-category')
@login_required
def spending_by_category():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    is_business = request.args.get('is_business', '0')
    conn = get_db_connection()
    vis_clause, vis_params = get_visible_clause('expenses', session['username'])
    rows = conn.execute(
        f"SELECT category, SUM(amount) as total FROM expenses WHERE strftime('%Y-%m',date)=? AND is_business=? AND {vis_clause} GROUP BY category ORDER BY total DESC",
        [month, int(is_business)] + vis_params).fetchall()
    conn.close()
    return jsonify({'data': [{'category': r[0], 'total': round(r[1],2)} for r in rows]}), 200

@app.route('/api/stats/monthly-trend')
@login_required
def monthly_trend():
    is_business = request.args.get('is_business', '0')
    conn = get_db_connection()
    exp_vis, exp_params = get_visible_clause('expenses', session['username'])
    inc_vis, inc_params = get_visible_clause('income', session['username'])
    rows = conn.execute(
        f"SELECT strftime('%Y-%m',date) as month, SUM(amount) as expenses FROM expenses WHERE is_business=? AND {exp_vis} GROUP BY month ORDER BY month DESC LIMIT 12",
        [int(is_business)] + exp_params
    ).fetchall()
    irows = conn.execute(
        f"SELECT strftime('%Y-%m',date) as month, SUM(amount) as income FROM income WHERE is_business=? AND {inc_vis} GROUP BY month ORDER BY month DESC LIMIT 12",
        [int(is_business)] + inc_params
    ).fetchall()
    conn.close()
    exp_map = {r[0]: round(r[1],2) for r in rows}
    inc_map = {r[0]: round(r[1],2) for r in irows}
    months = sorted(set(list(exp_map.keys()) + list(inc_map.keys())), reverse=True)[:12]
    data = [{'month': m, 'expenses': exp_map.get(m,0), 'income': inc_map.get(m,0)} for m in reversed(months)]
    return jsonify({'data': data}), 200

# ── Reports / Export ──────────────────────────────────────────────────────────

@app.route('/api/export/expenses/csv')
@login_required
def export_expenses_csv():
    start = request.args.get('start_date')
    end = request.args.get('end_date')
    conn = get_db_connection()
    q = "SELECT id,date,category,description,amount,paid_by,payment_method,receipt_filename,notes FROM expenses WHERE 1=1"
    params = []
    if start: q += " AND date >= ?"; params.append(start)
    if end: q += " AND date <= ?"; params.append(end)
    q += " ORDER BY date DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date','Category','Description','Amount','Paid By','Payment Method','Notes'])
    for r in rows:
        writer.writerow(list(r))
    output.seek(0)
    return app.response_class(output.getvalue(), mimetype='text/csv',
                               headers={'Content-Disposition': 'attachment;filename=expenses.csv'})

@app.route('/api/export/income/csv')
@login_required
def export_income_csv():
    conn = get_db_connection()
    rows = conn.execute("SELECT date,source,description,amount,received_by,notes FROM income ORDER BY date DESC").fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date','Source','Description','Amount','Received By','Notes'])
    for r in rows: writer.writerow(list(r))
    output.seek(0)
    return app.response_class(output.getvalue(), mimetype='text/csv',
                               headers={'Content-Disposition': 'attachment;filename=income.csv'})

@app.route('/api/export/expenses/excel')
@login_required
def export_expenses_excel():
    conn = get_db_connection()
    rows = conn.execute("SELECT date,category,description,amount,paid_by,payment_method,notes FROM expenses ORDER BY date DESC").fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    headers = ['Date','Category','Description','Amount','Paid By','Payment Method','Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(list(row), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return app.response_class(output.getvalue(),
                               mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               headers={'Content-Disposition': 'attachment;filename=expenses.xlsx'})

# ── Backup / Restore ──────────────────────────────────────────────────────────

@app.route('/api/backup', methods=['GET'])
@admin_required
def backup_data():
    conn = get_db_connection()
    data = {
        'version': '3.4',
        'exported_at': datetime.now().isoformat(),
        'expenses':     [dict(r) for r in conn.execute("SELECT * FROM expenses ORDER BY date").fetchall()],
        'income':       [dict(r) for r in conn.execute("SELECT * FROM income ORDER BY date").fetchall()],
        'credit_cards': [dict(r) for r in conn.execute("SELECT * FROM credit_cards").fetchall()],
        'budgets':      [dict(r) for r in conn.execute("SELECT * FROM budgets").fetchall()],
        'bills':        [dict(r) for r in conn.execute("SELECT * FROM bills").fetchall()],
        'recurring':    [dict(r) for r in conn.execute("SELECT * FROM recurring_transactions").fetchall()],
        'categories':   [dict(r) for r in conn.execute("SELECT * FROM categories").fetchall()],
    }
    conn.close()
    filename = f"kash-backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    return app.response_class(
        json.dumps(data, indent=2, default=str),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment;filename={filename}'}
    )

@app.route('/api/restore', methods=['POST'])
@admin_required
def restore_data():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.endswith('.json'):
        return jsonify({'error': 'File must be a .json backup file'}), 400
    try:
        data = json.loads(f.read().decode('utf-8'))
    except Exception:
        return jsonify({'error': 'Invalid JSON file'}), 400

    required = ['expenses','income','credit_cards','budgets','bills','recurring']
    if not all(k in data for k in required):
        return jsonify({'error': 'Invalid backup file — missing required sections'}), 400

    conn = get_db_connection()
    try:
        for table in ['expenses','income','credit_cards','budgets','bills','recurring_transactions']:
            conn.execute(f"DELETE FROM {table}")

        def insert_rows(table, rows, fields):
            for row in rows:
                vals = [row.get(f) for f in fields]
                placeholders = ','.join(['?']*len(fields))
                conn.execute(f"INSERT INTO {table} ({','.join(fields)}) VALUES ({placeholders})", vals)

        insert_rows('expenses', data['expenses'],
            ['date','category','description','amount','paid_by','payment_method','credit_card_id','notes','created_at'])
        insert_rows('income', data['income'],
            ['date','source','description','amount','received_by','notes','created_at'])
        insert_rows('credit_cards', data['credit_cards'],
            ['owner','card_name','credit_limit','manual_balance','created_at'])
        insert_rows('budgets', data['budgets'],
            ['category','amount','month','year'])
        insert_rows('bills', data['bills'],
            ['name','amount','due_date','category','is_paid','notes','created_at'])
        insert_rows('recurring_transactions', data['recurring'],
            ['type','frequency','next_date','description','amount','category','source','person',
             'payment_method','credit_card_id','is_active','created_at'])

        conn.commit()
        counts = {k: len(data[k]) for k in required}
        return jsonify({'message': 'Restore successful', 'counts': counts}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({'error': f'Restore failed: {str(e)}'}), 500
    finally:
        conn.close()


@app.route('/api/users/notifications', methods=['GET'])
@login_required
def get_notification_prefs():
    conn = get_db_connection()
    user = conn.execute(
        "SELECT email, notify_bills, notify_budgets, notify_monthly FROM users WHERE username=?",
        (session['username'],)).fetchone()
    conn.close()
    return jsonify({'data': dict(user)}), 200

@app.route('/api/users/notifications', methods=['PUT'])
@login_required
def update_notification_prefs():
    data = request.json or {}
    conn = get_db_connection()
    conn.execute("""UPDATE users SET email=?, notify_bills=?, notify_budgets=?, notify_monthly=?
                    WHERE username=?""",
                 (data.get('email','').strip(),
                  1 if data.get('notify_bills') else 0,
                  1 if data.get('notify_budgets') else 0,
                  1 if data.get('notify_monthly') else 0,
                  session['username']))
    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

@app.route('/api/notifications/test', methods=['POST'])
@login_required
def send_test_notification():
    if not app.config.get('MAIL_USERNAME'):
        return jsonify({'error': 'Email not configured on the server. Add MAIL_USERNAME and MAIL_PASSWORD to .env'}), 400
    conn = get_db_connection()
    user = conn.execute("SELECT email, display_name FROM users WHERE username=?", (session['username'],)).fetchone()
    conn.close()
    if not user or not user['email']:
        return jsonify({'error': 'Please save your email address first'}), 400
    html = f"""
    <div style="font-family:-apple-system,sans-serif;max-width:600px;margin:0 auto;">
      <div style="background:linear-gradient(135deg,#0a2540,#0d3d52);padding:24px 32px;border-radius:16px 16px 0 0;">
        <h1 style="margin:0;"><img src="https://raw.githubusercontent.com/ATW72/kash/52a2ed786417678f4be500380b41453cab4aaa27/static/email_logo.png" alt="Kash" height="24" style="vertical-align:middle;border:none;outline:none;"></h1>
      </div>
      <div style="background:#fff;padding:28px 32px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 16px 16px;">
        <p>Hi <strong>{user['display_name'] or session['username']}</strong>,</p>
        <p style="color:#10b981;font-weight:600;">✅ Your Kash notifications are working!</p>
        <p style="color:#6b7280;">You'll receive bill alerts, budget warnings, and monthly summaries at this address.</p>
      </div>
    </div>"""
    success = send_email(user['email'], '✅ Kash: Test Notification', html)
    if success:
        return jsonify({'success': True, 'message': f'Test email sent to {user["email"]}'}), 200
    return jsonify({'error': 'Failed to send email — check server mail configuration'}), 500

@app.route('/api/admin/mail-status', methods=['GET'])
@admin_required
def mail_status():
    return jsonify({
        'configured': bool(app.config.get('MAIL_USERNAME')),
        'server': app.config.get('MAIL_SERVER'),
        'username': app.config.get('MAIL_USERNAME', ''),
        'port': app.config.get('MAIL_PORT')
    }), 200

# ── Users ─────────────────────────────────────────────────────────────────────

@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    conn = get_db_connection()
    rows = [dict(r) for r in conn.execute("SELECT id,username,is_admin,display_name,email,created_at,must_change_password FROM users").fetchall()]
    conn.close()
    return jsonify({'data': rows}), 200



@app.route('/api/users/usernames')
@login_required
def get_usernames():
    conn = get_db_connection()
    rows = conn.execute("SELECT username FROM users ORDER BY username").fetchall()
    conn.close()
    return jsonify({'data': [r['username'] for r in rows]}), 200

@app.route('/api/users/profile', methods=['PUT'])
@login_required
def update_profile():
    data = request.json or {}
    display_name = data.get('display_name', '').strip()
    conn = get_db_connection()
    conn.execute("UPDATE users SET display_name=? WHERE username=?",
                 (display_name, session['username']))
    conn.commit()
    conn.close()
    session['display_name'] = display_name
    return jsonify({'success': True, 'display_name': display_name}), 200

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    data = request.json or {}
    username     = data.get('username', '').strip()
    display_name = data.get('display_name', '').strip()
    email        = data.get('email', '').strip()
    is_admin     = 1 if data.get('is_admin') else 0

    if not username:
        return jsonify({'error': 'Username is required'}), 400
    if not email:
        return jsonify({'error': 'Email is required to send the welcome email'}), 400

    # Auto-generate a secure temporary password
    alphabet = string.ascii_letters + string.digits
    temp_password = ''.join(secrets.choice(alphabet) for _ in range(12))
    phash = generate_password_hash(temp_password, method='pbkdf2:sha256')

    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""INSERT INTO users (username,password_hash,is_admin,display_name,email,must_change_password)
                     VALUES (?,?,?,?,?,1)""",
                  (username, phash, is_admin, display_name, email))
        conn.commit()
        conn.close()
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Username already exists'}), 400

    # Send welcome email with temp password
    # Use X-Forwarded headers if behind reverse proxy (Caddy), else fall back to host_url
    forwarded_host = request.headers.get('X-Forwarded-Host') or request.headers.get('Host', '')
    forwarded_proto = request.headers.get('X-Forwarded-Proto', 'https')
    if forwarded_host:
        app_url = f"{forwarded_proto}://{forwarded_host}"
    else:
        app_url = request.host_url.rstrip('/')
    html = build_welcome_email(username, display_name, temp_password, app_url)
    email_sent = send_email(email, '🎉 Welcome to Kash — Your Account is Ready', html)

    return jsonify({
        'success': True,
        'email_sent': email_sent,
        'message': f'User created and welcome email sent to {email}' if email_sent else f'User created but email failed — temp password: {temp_password}'
    }), 201


@app.route('/api/users/<int:uid>', methods=['PUT'])
@admin_required
def update_user(uid):
    data = request.json or {}
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'User not found'}), 404

    username     = data.get('username', row['username']).strip()
    display_name = data.get('display_name', row['display_name'] or '').strip()
    email        = data.get('email', row['email'] or '').strip()
    is_admin     = 1 if data.get('is_admin') else 0

    # Check username uniqueness if changed
    if username != row['username']:
        exists = conn.execute("SELECT id FROM users WHERE username=? AND id!=?", (username, uid)).fetchone()
        if exists:
            conn.close()
            return jsonify({'error': 'Username already taken'}), 400

    conn.execute("UPDATE users SET username=?,display_name=?,email=?,is_admin=? WHERE id=?",
                 (username, display_name, email, is_admin, uid))

    # Reset password if requested
    reset_sent = False
    if data.get('reset_password') and email:
        temp_password = ''.join(__import__('secrets').choice(__import__('string').ascii_letters + __import__('string').digits) for _ in range(12))
        phash = generate_password_hash(temp_password, method='pbkdf2:sha256')
        conn.execute("UPDATE users SET password_hash=?,must_change_password=1 WHERE id=?", (phash, uid))
        conn.commit()
        conn.close()
        forwarded_host = request.headers.get('X-Forwarded-Host') or request.headers.get('Host', '')
        forwarded_proto = request.headers.get('X-Forwarded-Proto', 'https')
        app_url2 = f"{forwarded_proto}://{forwarded_host}" if forwarded_host else request.host_url.rstrip('/')
        html = build_welcome_email(username, display_name, temp_password, app_url2)
        reset_sent = send_email(email, '🔐 Kash — Your Password Has Been Reset', html)
        return jsonify({'success': True, 'reset_sent': reset_sent}), 200

    conn.commit()
    conn.close()
    return jsonify({'success': True}), 200

@app.route('/api/users/<int:uid>', methods=['DELETE'])
@admin_required
def delete_user(uid):
    if uid == session.get('user_id'):
        return jsonify({'error': 'Cannot delete your own account'}), 400
    conn = get_db_connection()
    try:
        # First, check the user exists
        user = conn.execute("SELECT username FROM users WHERE id=?", (uid,)).fetchone()
        if not user:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
        username = user['username']
        # Clean up related data so FK constraints don't block the delete
        conn.execute("DELETE FROM sharing WHERE owner=? OR shared_with=?", (username, username))
        conn.execute("DELETE FROM audit_log WHERE username=?", (username,))
        conn.execute("DELETE FROM users WHERE id=?", (uid,))
        conn.commit()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.json or {}
    if not data.get('current_password') or not data.get('new_password'):
        return jsonify({'error': 'Current and new password required'}), 400
    if len(data['new_password']) < 8:
        return jsonify({'error': 'New password must be at least 8 characters'}), 400
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT password_hash FROM users WHERE id=?", (session['user_id'],))
    row = c.fetchone()
    if not row or not check_password_hash(row[0], data['current_password']):
        conn.close()
        return jsonify({'error': 'Current password is incorrect'}), 401
    phash = generate_password_hash(data['new_password'], method='pbkdf2:sha256')
    c.execute("UPDATE users SET password_hash=?,must_change_password=0 WHERE id=?", (phash, session['user_id']))
    conn.commit()
    conn.close()
    session['must_change_password'] = False
    return jsonify({'success': True, 'message': 'Password changed successfully'}), 200


@app.route('/api/users/force-change-password', methods=['POST'])
@login_required
def force_change_password():
    """Change password without requiring current password — only for must_change_password users."""
    if not session.get('must_change_password'):
        return jsonify({'error': 'Not required'}), 400
    data = request.json or {}
    if not data.get('new_password'):
        return jsonify({'error': 'New password required'}), 400
    if len(data['new_password']) < 8:
        return jsonify({'error': 'Password must be at least 8 characters'}), 400
    phash = generate_password_hash(data['new_password'], method='pbkdf2:sha256')
    conn = get_db_connection()
    conn.execute("UPDATE users SET password_hash=?,must_change_password=0 WHERE id=?",
                 (phash, session['user_id']))
    conn.commit()
    conn.close()
    session['must_change_password'] = False
    return jsonify({'success': True}), 200

@app.route('/health')
def health():
    return jsonify({'status': 'healthy', 'version': os.environ.get('APP_VERSION','3.4')}), 200

@app.route('/api/users/2fa/status', methods=['GET'])
@login_required
def get_2fa_status():
    conn = get_db_connection()
    row = conn.execute("SELECT two_factor_method FROM users WHERE id=?", (session['user_id'],)).fetchone()
    conn.close()
    return jsonify({'method': row['two_factor_method'] if row else 'none'}), 200

@app.route('/api/users/2fa/setup', methods=['POST'])
@login_required
def setup_2fa():
    """Generates a new TOTP secret or sends Email OTP for setup verification."""
    data = request.json or {}
    method = data.get('method', 'app')
    conn = get_db_connection()
    if method == 'app':
        import pyotp
        import qrcode
        import qrcode.image.svg
        import io
        secret = pyotp.random_base32()
        url = pyotp.totp.TOTP(secret).provisioning_uri(name=session['username'], issuer_name='Kash')
        # Store temporarily in session before confirming
        session['pending_2fa_secret'] = secret
        session['pending_2fa_method'] = method
        
        # Generate inline SVG string
        factory = qrcode.image.svg.SvgPathImage
        img = qrcode.make(url, image_factory=factory, border=0)
        stream = io.BytesIO()
        img.save(stream)
        svg_str = stream.getvalue().decode()
        
        conn.close()
        return jsonify({'qr_svg': svg_str, 'secret': secret}), 200
    elif method == 'email':
        email = conn.execute("SELECT email FROM users WHERE id=?", (session['user_id'],)).fetchone()[0]
        if not email:
            conn.close()
            return jsonify({'error': 'Email not configured on your profile.'}), 400
        import random
        otp = f"{random.randint(100000, 999999)}"
        session['pending_2fa_secret'] = otp # Use secret for email OTP as well
        session['pending_2fa_method'] = method
        from utils.notifications import send_email
        send_email(email, '✅ Kash: 2FA Setup Code', f"<h2>2FA Verification Code</h2><p>Your setup code is: <strong>{otp}</strong></p><p>If you didn't request this, you can safely ignore this email.</p>")
        conn.close()
        return jsonify({'message': 'Check your email for the code.'}), 200
    conn.close()
    return jsonify({'error': 'Invalid 2FA method'}), 400

@app.route('/api/users/2fa/verify', methods=['POST'])
@login_required
def verify_2fa_setup():
    data = request.json or {}
    code = data.get('code', '').strip()
    pending_method = session.get('pending_2fa_method')
    pending_secret = session.get('pending_2fa_secret')
    
    if not pending_method or not pending_secret:
        return jsonify({'error': 'No 2FA setup in progress.'}), 400
        
    conn = get_db_connection()
    if pending_method == 'app':
        import pyotp
        totp = pyotp.TOTP(pending_secret)
        if not totp.verify(code):
            conn.close()
            return jsonify({'error': 'Invalid Authenticator code. Please try again.'}), 400
        conn.execute("UPDATE users SET two_factor_method='app', totp_secret=? WHERE id=?", (pending_secret, session['user_id']))
    elif pending_method == 'email':
        if code != str(pending_secret):
            conn.close()
            return jsonify({'error': 'Invalid Email code.'}), 400
        conn.execute("UPDATE users SET two_factor_method='email', totp_secret='' WHERE id=?", (session['user_id'],))
        
    conn.commit()
    conn.close()
    session.pop('pending_2fa_method', None)
    session.pop('pending_2fa_secret', None)
    return jsonify({'success': True, 'message': '2FA enabled successfully!'}), 200

@app.route('/api/users/2fa/disable', methods=['POST'])
@login_required
def disable_2fa():
    conn = get_db_connection()
    conn.execute("UPDATE users SET two_factor_method='none', totp_secret='' WHERE id=?", (session['user_id'],))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': '2FA has been disabled.'}), 200

if __name__ == '__main__':
    init_db()
    app.run(host=Config.HOST, port=Config.PORT, debug=Config.DEBUG)
